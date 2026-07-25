"""Build and validate provisional evidence resources for QBI v3.

This module owns deterministic estimation logic.  The command-line tool in
``tools/build_us_qbi_v3_evidence.py`` owns restricted-file I/O, input hashing,
and writing the derived JSON resources.  Nothing in this module wires the
evidence into a simulation.
"""

from __future__ import annotations

import math
import re
from collections.abc import Mapping, Sequence
from dataclasses import dataclass
from pathlib import Path

import numpy as np
import pandas as pd

SCF_IMPLICATE_COUNT = 5
SCF_MINIMUM_UNWEIGHTED_N = 30.0
SCF_MARGIN_QUANTILES = (0.05, 0.25, 0.5, 0.75, 0.95)
JCT_ZERO_EMPLOYEE_FIRM_SHARE = 0.842

SCF_REQUIRED_COLUMNS = (
    "y1",
    "x42001",
    "x3103",
    "x3104",
    "x3105",
    "x3107",
    "x3111",
    "x3113",
    "x3114",
    "x3119",
    "x3128",
    "x3131",
    "x3132",
    "x3207",
    "x3211",
    "x3213",
    "x3214",
    "x3219",
    "x3228",
    "x3231",
    "x3232",
)

SCF_INCOME_BANDS = (
    "nonpositive",
    "0_to_25k",
    "25k_to_100k",
    "100k_to_250k",
    "250k_to_1m",
    "over_1m",
)
SCF_LEGAL_FORM_GROUPS = (
    "partnership_or_llc",
    "sole_or_informal",
    "s_corporation",
    "other_or_unknown",
)
SCF_INDUSTRY_BINS = (1, 2, 3, 4, 5, 6, 7, 99)
SCF_HEADCOUNT_SIZE_BANDS = (
    "2_to_4",
    "5_to_9",
    "10_to_24",
    "25_to_99",
    "100_plus",
)

_LEGAL_FORM_BY_CODE = {
    1: "partnership_or_llc",
    11: "partnership_or_llc",
    2: "sole_or_informal",
    40: "sole_or_informal",
    3: "s_corporation",
    4: "other_or_unknown",
    -7: "other_or_unknown",
}
_MAIN_QBI_PROXY_CODES = frozenset({1, 2, 3, 11, 40})
_STRICT_QBI_PROXY_CODES = frozenset({1, 2, 3, 11})

SOI_ENTITY_FORMS = (
    "sole_proprietorship",
    "partnership",
    "s_corporation",
)
SOI_PUBLICATION_FLAGS = frozenset(
    {
        "published",
        "caution",
        "combined_for_disclosure",
        "deleted_for_disclosure",
        "not_published",
    }
)


@dataclass(frozen=True)
class _EstimateSelection:
    """The sample selected by a thin-cell fallback hierarchy."""

    level: str
    frame: pd.DataFrame
    dimensions: Mapping[str, object]


def _finite_numeric(frame: pd.DataFrame, columns: Sequence[str]) -> pd.DataFrame:
    result = frame.copy()
    for column in columns:
        result[column] = pd.to_numeric(result[column], errors="coerce")
        invalid = ~np.isfinite(result[column].to_numpy(dtype=np.float64))
        if invalid.any():
            raise ValueError(
                f"SCF column {column!r} contains {int(invalid.sum())} "
                "nonnumeric or nonfinite value(s)."
            )
    return result


def _normalized_scf_frame(frame: pd.DataFrame) -> pd.DataFrame:
    normalized = frame.rename(
        columns={str(column): str(column).lower() for column in frame}
    )
    duplicate_columns = normalized.columns[normalized.columns.duplicated()].tolist()
    if duplicate_columns:
        raise ValueError(
            "SCF frame has case-insensitive duplicate columns "
            f"{sorted(set(duplicate_columns))}."
        )
    missing = sorted(set(SCF_REQUIRED_COLUMNS) - set(normalized.columns))
    if missing:
        raise ValueError(f"SCF frame is missing required columns {missing}.")
    return _finite_numeric(normalized, SCF_REQUIRED_COLUMNS)


def _income_band(values: pd.Series) -> pd.Series:
    return pd.cut(
        values,
        bins=(-np.inf, 0.0, 25_000.0, 100_000.0, 250_000.0, 1_000_000.0, np.inf),
        labels=SCF_INCOME_BANDS,
        right=True,
        include_lowest=True,
    ).astype("object")


def _headcount_size_band(values: pd.Series) -> pd.Series:
    return pd.cut(
        values,
        bins=(1.0, 4.0, 9.0, 24.0, 99.0, np.inf),
        labels=SCF_HEADCOUNT_SIZE_BANDS,
        right=True,
    ).astype("object")


def build_scf_business_records(frame: pd.DataFrame) -> pd.DataFrame:
    """Stack the first two actively managed SCF business records.

    Each returned row is one household-business-implicate record.  The point
    estimate weight is ``X42001 / 5``; ``unweighted_n`` elsewhere in this
    module similarly divides pooled record counts by five so the implicates
    are not treated as independent observations.
    """

    source = _normalized_scf_frame(frame)
    y1 = source["y1"].to_numpy(dtype=np.int64)
    if not np.array_equal(y1.astype(np.float64), source["y1"].to_numpy()):
        raise ValueError("SCF Y1 must contain integer identifiers.")
    if pd.Series(y1).duplicated().any():
        raise ValueError("SCF Y1 must be unique across household implicates.")
    household_id = y1 // 10
    implicate = y1 % 10
    expected_implicates = set(range(1, SCF_IMPLICATE_COUNT + 1))
    by_household = pd.DataFrame(
        {"household_id": household_id, "implicate": implicate}
    ).groupby("household_id")["implicate"]
    malformed = [
        int(identifier)
        for identifier, values in by_household
        if set(values.tolist()) != expected_implicates
    ]
    if malformed:
        raise ValueError(
            "SCF Y1 must encode exactly implicates 1-5 for every household; "
            f"malformed household ids include {malformed[:5]}."
        )

    active = source["x3103"].eq(1) & source["x3104"].eq(1) & source["x3105"].ge(1)
    pooled_weight = source["x42001"] / SCF_IMPLICATE_COUNT
    if (pooled_weight <= 0.0).any():
        raise ValueError("SCF X42001 must be positive for every implicate.")

    records: list[pd.DataFrame] = []
    slot_columns = {
        1: {
            "industry_code": "x3107",
            "headcount": "x3111",
            "respondent_works": "x3113",
            "spouse_works": "x3114",
            "legal_form_code": "x3119",
            "ownership_percent_x100": "x3128",
            "gross_receipts": "x3131",
            "whole_net_income": "x3132",
        },
        2: {
            "industry_code": "x3207",
            "headcount": "x3211",
            "respondent_works": "x3213",
            "spouse_works": "x3214",
            "legal_form_code": "x3219",
            "ownership_percent_x100": "x3228",
            "gross_receipts": "x3231",
            "whole_net_income": "x3232",
        },
    }
    for slot, columns in slot_columns.items():
        selected = active & source["x3105"].ge(slot)
        slot_frame = source.loc[selected, list(columns.values())].rename(
            columns={value: key for key, value in columns.items()}
        )
        slot_frame.insert(0, "y1", y1[selected])
        slot_frame.insert(1, "household_id", household_id[selected])
        slot_frame.insert(2, "implicate", implicate[selected])
        slot_frame.insert(3, "business_slot", slot)
        slot_frame.insert(4, "weight", pooled_weight.loc[selected].to_numpy())
        records.append(slot_frame.reset_index(drop=True))

    if not records:
        raise ValueError("SCF frame produced no actively managed business records.")
    business = pd.concat(records, ignore_index=True)
    if business.empty:
        raise ValueError("SCF frame produced no actively managed business records.")

    integer_columns = (
        "industry_code",
        "headcount",
        "respondent_works",
        "spouse_works",
        "legal_form_code",
        "ownership_percent_x100",
    )
    for column in integer_columns:
        values = business[column].to_numpy(dtype=np.float64)
        if not np.array_equal(values, values.astype(np.int64)):
            raise ValueError(f"SCF business field {column!r} must be integer-coded.")
        business[column] = values.astype(np.int64)

    business["headcount"] = business["headcount"].replace(-1, 0)
    if (business["headcount"] < 0).any():
        raise ValueError(
            "SCF business headcount contains an unsupported negative code."
        )
    business["gross_receipts"] = business["gross_receipts"].replace(-1.0, 0.0)
    business["whole_net_income"] = business["whole_net_income"].replace(-1.0, 0.0)
    if (business["gross_receipts"] < 0.0).any():
        raise ValueError("SCF gross receipts contain an unsupported negative value.")
    ownership = business["ownership_percent_x100"]
    if ((ownership < 0) | (ownership > 10_000)).any():
        raise ValueError("SCF ownership percentage must lie between 0 and 10,000.")

    unsupported_forms = sorted(
        set(business["legal_form_code"]) - set(_LEGAL_FORM_BY_CODE)
    )
    if unsupported_forms:
        raise ValueError(
            f"SCF business records contain unsupported legal-form codes "
            f"{unsupported_forms}."
        )
    unsupported_industries = sorted(
        set(business["industry_code"]) - set(SCF_INDUSTRY_BINS)
    )
    if unsupported_industries:
        raise ValueError(
            f"SCF business records contain unsupported industry codes "
            f"{unsupported_industries}."
        )

    business["legal_form_group"] = business["legal_form_code"].map(_LEGAL_FORM_BY_CODE)
    business["owned_net_income"] = business["whole_net_income"] * ownership / 10_000.0
    business["income_band"] = _income_band(business["owned_net_income"])
    business["employer_presence_proxy"] = business["headcount"] > 1
    business["headcount_size_band"] = _headcount_size_band(business["headcount"])
    business["respondent_or_spouse_works"] = business["respondent_works"].eq(
        1
    ) | business["spouse_works"].eq(1)
    business["qbi_positive_proxy"] = business["owned_net_income"].gt(0.0) & business[
        "legal_form_code"
    ].isin(_MAIN_QBI_PROXY_CODES)
    business["qbi_positive_strict_form_proxy"] = business["owned_net_income"].gt(
        0.0
    ) & business["legal_form_code"].isin(_STRICT_QBI_PROXY_CODES)
    return business.reset_index(drop=True)


def weighted_inverse_cdf(
    values: Sequence[float] | np.ndarray,
    weights: Sequence[float] | np.ndarray,
    probabilities: Sequence[float] = SCF_MARGIN_QUANTILES,
) -> dict[str, float]:
    """Return weighted inverse-CDF quantiles with stable tie handling."""

    value_array = np.asarray(values, dtype=np.float64)
    weight_array = np.asarray(weights, dtype=np.float64)
    probability_array = np.asarray(probabilities, dtype=np.float64)
    if value_array.shape != weight_array.shape:
        raise ValueError("Weighted quantile values and weights must have equal shape.")
    if value_array.ndim != 1:
        raise ValueError(
            "Weighted quantile values and weights must be one-dimensional."
        )
    if (
        not np.isfinite(value_array).all()
        or not np.isfinite(weight_array).all()
        or (weight_array <= 0.0).any()
    ):
        raise ValueError(
            "Weighted quantiles require finite values and positive weights."
        )
    if value_array.size == 0:
        raise ValueError("Weighted quantiles require at least one observation.")
    if (
        not np.isfinite(probability_array).all()
        or (probability_array < 0.0).any()
        or (probability_array > 1.0).any()
    ):
        raise ValueError("Weighted quantile probabilities must lie in [0, 1].")

    order = np.argsort(value_array, kind="stable")
    sorted_values = value_array[order]
    cumulative = np.cumsum(weight_array[order])
    targets = probability_array * cumulative[-1]
    positions = np.searchsorted(cumulative, targets, side="left")
    positions = np.minimum(positions, len(sorted_values) - 1)
    return {
        _quantile_name(probability): float(sorted_values[position])
        for probability, position in zip(probability_array, positions, strict=True)
    }


def _quantile_name(probability: float) -> str:
    return f"q{int(round(probability * 100)):02d}"


def _unweighted_n(frame: pd.DataFrame) -> float:
    return float(len(frame) / SCF_IMPLICATE_COUNT)


def _counts(frame: pd.DataFrame) -> dict[str, float | int]:
    employer = frame.loc[frame["employer_presence_proxy"]]
    return {
        "pooled_record_count": int(len(frame)),
        "implicate_adjusted_unweighted_n": _unweighted_n(frame),
        "weighted_business_interests": float(frame["weight"].sum()),
        "employer_proxy_pooled_record_count": int(len(employer)),
        "employer_proxy_implicate_adjusted_unweighted_n": _unweighted_n(employer),
        "weighted_employer_proxy_business_interests": float(employer["weight"].sum()),
    }


def _candidate_sample(
    records: pd.DataFrame,
    *,
    income_band: str,
    legal_form_group: str,
    industry_code: int,
    employer_only: bool,
    minimum_n: float,
) -> _EstimateSelection:
    base = (
        records["employer_presence_proxy"]
        if employer_only
        else pd.Series(True, index=records.index)
    )
    candidates = (
        (
            "exact",
            base
            & records["income_band"].eq(income_band)
            & records["legal_form_group"].eq(legal_form_group)
            & records["industry_code"].eq(industry_code),
            {
                "income_band": income_band,
                "legal_form_group": legal_form_group,
                "industry_code": industry_code,
            },
        ),
        (
            "income_form",
            base
            & records["income_band"].eq(income_band)
            & records["legal_form_group"].eq(legal_form_group),
            {
                "income_band": income_band,
                "legal_form_group": legal_form_group,
                "industry_code": "all",
            },
        ),
        (
            "form",
            base & records["legal_form_group"].eq(legal_form_group),
            {
                "income_band": "all",
                "legal_form_group": legal_form_group,
                "industry_code": "all",
            },
        ),
        (
            "all",
            base,
            {
                "income_band": "all",
                "legal_form_group": "all",
                "industry_code": "all",
            },
        ),
    )
    for index, (level, mask, dimensions) in enumerate(candidates):
        sample = records.loc[mask]
        if _unweighted_n(sample) >= minimum_n or index == len(candidates) - 1:
            if sample.empty:
                raise ValueError(
                    "SCF fallback hierarchy reached an empty global sample."
                )
            return _EstimateSelection(level, sample, dimensions)
    raise AssertionError("SCF fallback hierarchy did not select a sample.")


def _presence_estimate(selection: _EstimateSelection) -> dict[str, object]:
    weights = selection.frame["weight"].to_numpy(dtype=np.float64)
    present = selection.frame["employer_presence_proxy"].to_numpy(dtype=np.float64)
    return {
        "probability_headcount_gt_1": float(np.sum(weights * present) / weights.sum()),
        "estimate_level": selection.level,
        "source_dimensions": dict(selection.dimensions),
        "source_counts": _counts(selection.frame),
    }


def _size_estimate(selection: _EstimateSelection) -> dict[str, object]:
    denominator = float(selection.frame["weight"].sum())
    shares = {}
    for band in SCF_HEADCOUNT_SIZE_BANDS:
        weight = selection.frame.loc[
            selection.frame["headcount_size_band"].eq(band), "weight"
        ].sum()
        shares[band] = float(weight / denominator)
    return {
        "conditional_on": "headcount_gt_1",
        "shares": shares,
        "estimate_level": selection.level,
        "source_dimensions": dict(selection.dimensions),
        "source_counts": _counts(selection.frame),
    }


def _comparison(
    records: pd.DataFrame,
    flag: str,
    *,
    legal_form_codes: frozenset[int],
) -> dict[str, object]:
    sample = records.loc[records[flag]]
    if sample.empty:
        raise ValueError(f"SCF comparison flag {flag!r} selected no records.")
    weights = sample["weight"].to_numpy(dtype=np.float64)
    zero_proxy = ~sample["employer_presence_proxy"].to_numpy(dtype=bool)
    share = float(weights[zero_proxy].sum() / weights.sum())
    return {
        "sample_definition": (
            "Owned net income is greater than zero and the public legal-form "
            f"code is in {sorted(legal_form_codes)}. The denominator is pooled "
            "household-business-implicate interests weighted by X42001/5."
        ),
        "numerator_definition": (
            "The denominator sample whose SCF headcount is at most one."
        ),
        "legal_form_codes": sorted(legal_form_codes),
        "pooled_record_count": int(len(sample)),
        "implicate_adjusted_unweighted_n": _unweighted_n(sample),
        "weighted_business_interests": float(weights.sum()),
        "headcount_le_1_share": share,
        "difference_from_jct_percentage_points": float(
            100.0 * (share - JCT_ZERO_EMPLOYEE_FIRM_SHARE)
        ),
    }


def _profit_margin_cells(
    records: pd.DataFrame,
    *,
    minimum_n: float,
) -> list[dict[str, object]]:
    positive = records.loc[
        records["owned_net_income"].gt(0.0) & records["gross_receipts"].gt(0.0)
    ].copy()
    positive["profit_margin"] = (
        positive["whole_net_income"] / positive["gross_receipts"]
    )
    if positive.empty:
        raise ValueError("SCF records contain no positive profit-margin sample.")

    cells: list[dict[str, object]] = []
    for legal_form_group in SCF_LEGAL_FORM_GROUPS:
        form_sample = positive.loc[positive["legal_form_group"].eq(legal_form_group)]
        for industry_code in SCF_INDUSTRY_BINS:
            exact = form_sample.loc[form_sample["industry_code"].eq(industry_code)]
            candidates = (
                (
                    "exact",
                    exact,
                    {
                        "legal_form_group": legal_form_group,
                        "industry_code": industry_code,
                    },
                ),
                (
                    "form",
                    form_sample,
                    {
                        "legal_form_group": legal_form_group,
                        "industry_code": "all",
                    },
                ),
                (
                    "all",
                    positive,
                    {"legal_form_group": "all", "industry_code": "all"},
                ),
            )
            selection: _EstimateSelection | None = None
            for index, (level, sample, dimensions) in enumerate(candidates):
                if _unweighted_n(sample) >= minimum_n or index == len(candidates) - 1:
                    selection = _EstimateSelection(level, sample, dimensions)
                    break
            if selection is None or selection.frame.empty:
                raise ValueError("SCF profit-margin fallback selected no records.")
            margins = selection.frame["profit_margin"].to_numpy(dtype=np.float64)
            weights = selection.frame["weight"].to_numpy(dtype=np.float64)
            cells.append(
                {
                    "legal_form_group": legal_form_group,
                    "industry_code": industry_code,
                    "requested_counts": {
                        "pooled_record_count": int(len(exact)),
                        "implicate_adjusted_unweighted_n": _unweighted_n(exact),
                        "weighted_business_interests": float(exact["weight"].sum()),
                    },
                    "estimate_level": selection.level,
                    "source_dimensions": dict(selection.dimensions),
                    "source_counts": {
                        "pooled_record_count": int(len(selection.frame)),
                        "implicate_adjusted_unweighted_n": _unweighted_n(
                            selection.frame
                        ),
                        "weighted_business_interests": float(weights.sum()),
                    },
                    "quantiles": weighted_inverse_cdf(
                        margins, weights, SCF_MARGIN_QUANTILES
                    ),
                    "source_minimum": float(margins.min()),
                    "source_maximum": float(margins.max()),
                }
            )
    return cells


def build_qbi_employer_structure_resource(
    frame: pd.DataFrame,
    *,
    provenance: Mapping[str, object],
    minimum_unweighted_n: float = SCF_MINIMUM_UNWEIGHTED_N,
) -> dict[str, object]:
    """Build the provisional SCF employer-structure evidence resource."""

    if not math.isfinite(minimum_unweighted_n) or minimum_unweighted_n <= 0.0:
        raise ValueError("SCF minimum unweighted n must be positive and finite.")
    records = build_scf_business_records(frame)
    cells: list[dict[str, object]] = []
    for income_band in SCF_INCOME_BANDS:
        for legal_form_group in SCF_LEGAL_FORM_GROUPS:
            for industry_code in SCF_INDUSTRY_BINS:
                exact = records.loc[
                    records["income_band"].eq(income_band)
                    & records["legal_form_group"].eq(legal_form_group)
                    & records["industry_code"].eq(industry_code)
                ]
                presence = _candidate_sample(
                    records,
                    income_band=income_band,
                    legal_form_group=legal_form_group,
                    industry_code=industry_code,
                    employer_only=False,
                    minimum_n=minimum_unweighted_n,
                )
                size = _candidate_sample(
                    records,
                    income_band=income_band,
                    legal_form_group=legal_form_group,
                    industry_code=industry_code,
                    employer_only=True,
                    minimum_n=minimum_unweighted_n,
                )
                cells.append(
                    {
                        "income_band": income_band,
                        "legal_form_group": legal_form_group,
                        "industry_code": industry_code,
                        "requested_counts": _counts(exact),
                        "employer_presence": _presence_estimate(presence),
                        "headcount_size_distribution": _size_estimate(size),
                    }
                )

    payload: dict[str, object] = {
        "schema_version": 1,
        "resource_id": "qbi_employer_structure_v1",
        "provisional": True,
        "source": {
            "survey": "2022 Survey of Consumer Finances full public data",
            "survey_year": 2022,
            "business_flow_year": 2021,
            "record_scope": (
                "First and second detailed actively managed household businesses"
            ),
            "record_selection": (
                "Require X3103 == 1, X3104 == 1, and X3105 >= the requested "
                "business slot (1 or 2). X3113/X3114 and X3213/X3214 describe "
                "whether the respondent or spouse works in the business but do "
                "not further restrict the family-level actively managed sample."
            ),
            "variables": {
                "active_management_screeners": ["X3103", "X3104"],
                "business_count": "X3105",
                "weight": "X42001",
                "industry": ["X3107", "X3207"],
                "headcount": ["X3111", "X3211"],
                "legal_form": ["X3119", "X3219"],
                "ownership_percent_x100": ["X3128", "X3228"],
                "gross_receipts": ["X3131", "X3231"],
                "whole_business_net_income": ["X3132", "X3232"],
                "respondent_or_spouse_works": [
                    "X3113",
                    "X3114",
                    "X3213",
                    "X3214",
                ],
            },
            "codebook": {
                "url": ("https://www.federalreserve.gov/econres/files/codebk2022.txt"),
                "implicate_convention_lines": "622-655",
                "weight_definition_lines": "1865-1873, 2792-2796",
                "business_variable_lines": "11699-12534",
            },
        },
        "methodology": {
            "point_estimate_weight": "X42001 divided by 5",
            "implicate_count": SCF_IMPLICATE_COUNT,
            "implicate_note": (
                "Point estimates pool all five implicates with record weight "
                "X42001/5, following the SCF simple-statistic convention. "
                "No replicate-weight or multiple-imputation sampling-error "
                "estimate is claimed, and implicates are not treated as "
                "independent observations."
            ),
            "income_measure": (
                "Whole-business pretax net income multiplied by the household "
                "ownership percentage; exact code -1 ('Nothing') is zero."
            ),
            "employer_presence_proxy": "headcount greater than 1",
            "employer_proxy_limitation": (
                "SCF headcount includes owners, family members, unpaid workers, "
                "and all paid full- and part-time workers, so headcount>1 is an "
                "upper-bound employer-presence proxy."
            ),
            "minimum_implicate_adjusted_unweighted_n": minimum_unweighted_n,
            "presence_collapse_order": [
                "income_form_industry",
                "income_form",
                "form",
                "all",
            ],
            "size_collapse_order": [
                "income_form_industry_among_headcount_gt_1",
                "income_form_among_headcount_gt_1",
                "form_among_headcount_gt_1",
                "all_among_headcount_gt_1",
            ],
            "size_collapse_note": (
                "Employer size is collapsed independently on its headcount>1 "
                "denominator so a cell with 30 total businesses but few employer "
                "proxies does not retain a thin size distribution."
            ),
            "headcount_disclosure": (
                "Counts above 10 are rounded to the nearest 5 and top-coded at 5,000."
            ),
        },
        "dimensions": {
            "income_bands": [
                {"id": "nonpositive", "definition": "owned net income <= 0"},
                {"id": "0_to_25k", "definition": "0 < owned net income <= 25,000"},
                {
                    "id": "25k_to_100k",
                    "definition": "25,000 < owned net income <= 100,000",
                },
                {
                    "id": "100k_to_250k",
                    "definition": "100,000 < owned net income <= 250,000",
                },
                {
                    "id": "250k_to_1m",
                    "definition": "250,000 < owned net income <= 1,000,000",
                },
                {"id": "over_1m", "definition": "owned net income > 1,000,000"},
            ],
            "legal_form_groups": [
                {"id": "partnership_or_llc", "public_codes": [1, 11]},
                {"id": "sole_or_informal", "public_codes": [2, 40]},
                {"id": "s_corporation", "public_codes": [3]},
                {"id": "other_or_unknown", "public_codes": [4, -7]},
            ],
            "industry_bins": [
                {"code": 1, "summary": "agriculture plus veterinary/landscaping"},
                {"code": 2, "summary": "mining and construction"},
                {"code": 3, "summary": "manufacturing plus selected publishing"},
                {"code": 4, "summary": "wholesale/retail plus food services"},
                {
                    "code": 5,
                    "summary": (
                        "finance/real estate plus selected information, "
                        "business support, rental, and repair"
                    ),
                },
                {"code": 6, "summary": "remaining private service industries"},
                {"code": 7, "summary": "public administration and military"},
                {
                    "code": 99,
                    "summary": (
                        "industry suppressed for very-high-value businesses; "
                        "not an economic industry"
                    ),
                },
            ],
            "headcount_size_bands": [
                {"id": "2_to_4", "definition": "2-4 people"},
                {"id": "5_to_9", "definition": "5-9 people"},
                {"id": "10_to_24", "definition": "10-24 people"},
                {"id": "25_to_99", "definition": "25-99 people"},
                {"id": "100_plus", "definition": "100 or more people"},
            ],
        },
        "external_anchor": {
            "source": (
                "Joint Committee on Taxation, Data Related to Deductions "
                "Claimed Under Code Section 199A, Table 5"
            ),
            "source_url": (
                "https://www.warren.senate.gov/imo/media/doc/jct_report_on_199a.pdf"
            ),
            "tax_year": 2022,
            "zero_employee_firm_share": JCT_ZERO_EMPLOYEE_FIRM_SHARE,
            "zero_employee_deduction_dollar_share": 0.357,
            "sole_proprietor_zero_employee_share_statement": "more than 95%",
            "partnership_zero_employee_share_statement": "more than 80%",
            "scf_comparison": {
                "main_proxy_including_informal_code_40": _comparison(
                    records,
                    "qbi_positive_proxy",
                    legal_form_codes=_MAIN_QBI_PROXY_CODES,
                ),
                "strict_form_sensitivity_excluding_code_40": _comparison(
                    records,
                    "qbi_positive_strict_form_proxy",
                    legal_form_codes=_STRICT_QBI_PROXY_CODES,
                ),
            },
            "definition_gaps": [
                (
                    "JCT counts deduction-generating tax firms with zero W-2 "
                    "employees; SCF observes household business interests and "
                    "only whether reported headcount exceeds one."
                ),
                (
                    "SCF headcount includes owners, family, unpaid workers, and "
                    "paid workers, making its headcount<=1 share a lower bound "
                    "on a zero-W-2-employer share."
                ),
                (
                    "JCT uses tax year 2022 and actual QBID-generating firms; "
                    "SCF survey-year 2022 reports calendar-2021 pretax income "
                    "without SSTB, aggregation, carryforward, taxable-income, "
                    "or deduction eligibility rules."
                ),
                (
                    "SCF details only the first/largest two actively managed "
                    "businesses and cannot deduplicate a firm owned by households "
                    "observed separately."
                ),
            ],
        },
        "cells": cells,
        "profit_margin_quantiles": {
            "definition": (
                "Whole-business pretax net income divided by whole-business "
                "gross receipts, among positive owned net income and receipts."
            ),
            "ownership_note": (
                "The ownership percentage cancels from the numerator and "
                "denominator of the margin."
            ),
            "support_note": (
                "Margins are empirical and uncapped; SCF permits reported net "
                "income above gross receipts."
            ),
            "quantile_method": (
                "weighted inverse CDF: smallest margin whose cumulative "
                "X42001/5 weight reaches the requested probability"
            ),
            "probabilities": list(SCF_MARGIN_QUANTILES),
            "minimum_implicate_adjusted_unweighted_n": minimum_unweighted_n,
            "collapse_order": ["form_industry", "form", "all"],
            "cells": _profit_margin_cells(records, minimum_n=minimum_unweighted_n),
        },
        "provenance": dict(provenance),
        "judgment_calls": [
            (
                "Income bands use the household-owned share of whole-business "
                "net income to align with recipient QBI; whole-business income "
                "would be a more direct predictor of firm headcount."
            ),
            (
                "Code 40 ('not a formal business type') is grouped with sole "
                "proprietorships in the main comparison; a strict sensitivity "
                "excludes it."
            ),
            (
                "The actively managed family-business screen does not require "
                "the respondent or spouse to be the working family member; "
                "X3113/X3114 and X3213/X3214 are retained only as diagnostics."
            ),
            (
                "Band boundaries, minimum n=30, the nested collapse hierarchy, "
                "independent employer-size collapse, and the five-point margin "
                "quantile grid are modeling choices pending supervisor review."
            ),
        ],
    }
    validate_qbi_employer_structure_resource(payload)
    return payload


def _mapping(value: object, location: str) -> Mapping[str, object]:
    if not isinstance(value, Mapping):
        raise ValueError(f"{location} must be a JSON object.")
    return value


def _list(value: object, location: str) -> list[object]:
    if not isinstance(value, list):
        raise ValueError(f"{location} must be a JSON array.")
    return value


def _finite_probability(value: object, location: str) -> float:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise ValueError(f"{location} must be numeric.")
    result = float(value)
    if not math.isfinite(result) or not 0.0 <= result <= 1.0:
        raise ValueError(f"{location} must be finite and lie in [0, 1].")
    return result


def _exact_mapping(
    value: object,
    expected_keys: set[str] | frozenset[str],
    location: str,
) -> Mapping[str, object]:
    result = _mapping(value, location)
    if set(result) != set(expected_keys):
        raise ValueError(
            f"{location} keys must be exactly {sorted(expected_keys)}; "
            f"got {sorted(result)}."
        )
    return result


def _nonempty_text(value: object, location: str) -> str:
    if not isinstance(value, str) or not value.strip():
        raise ValueError(f"{location} must be a nonempty string.")
    return value


def _finite_number(
    value: object,
    location: str,
    *,
    minimum: float | None = None,
) -> float:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise ValueError(f"{location} must be numeric.")
    result = float(value)
    if not math.isfinite(result) or (minimum is not None and result < minimum):
        floor = "" if minimum is None else f" and at least {minimum}"
        raise ValueError(f"{location} must be finite{floor}.")
    return result


def _nonnegative_integer(value: object, location: str) -> int:
    if isinstance(value, bool) or not isinstance(value, int) or value < 0:
        raise ValueError(f"{location} must be a nonnegative integer.")
    return value


def _validate_text_list(value: object, location: str) -> list[object]:
    result = _list(value, location)
    if not result:
        raise ValueError(f"{location} must be nonempty.")
    for index, item in enumerate(result):
        _nonempty_text(item, f"{location}[{index}]")
    return result


def _validate_provenance(value: object, location: str) -> None:
    provenance = _mapping(value, location)
    _nonempty_text(provenance.get("generated_by"), f"{location}.generated_by")
    _nonempty_text(provenance.get("run_command"), f"{location}.run_command")
    inputs = _list(provenance.get("inputs"), f"{location}.inputs")
    if not inputs:
        raise ValueError(f"{location}.inputs must be nonempty.")
    for index, raw_input in enumerate(inputs):
        input_record = _mapping(raw_input, f"{location}.inputs[{index}]")
        _nonempty_text(
            input_record.get("filename"), f"{location}.inputs[{index}].filename"
        )
        digest = _nonempty_text(
            input_record.get("sha256"), f"{location}.inputs[{index}].sha256"
        )
        if re.fullmatch(r"[0-9a-f]{64}", digest) is None:
            raise ValueError(
                f"{location}.inputs[{index}].sha256 must be lowercase SHA-256."
            )


_SCF_COUNT_KEYS = frozenset(
    {
        "pooled_record_count",
        "implicate_adjusted_unweighted_n",
        "weighted_business_interests",
        "employer_proxy_pooled_record_count",
        "employer_proxy_implicate_adjusted_unweighted_n",
        "weighted_employer_proxy_business_interests",
    }
)
_SCF_SIMPLE_COUNT_KEYS = frozenset(
    {
        "pooled_record_count",
        "implicate_adjusted_unweighted_n",
        "weighted_business_interests",
    }
)


def _validate_scf_counts(
    value: object,
    location: str,
    *,
    include_employer: bool,
) -> Mapping[str, object]:
    expected = _SCF_COUNT_KEYS if include_employer else _SCF_SIMPLE_COUNT_KEYS
    counts = _exact_mapping(value, expected, location)
    pooled = _nonnegative_integer(
        counts["pooled_record_count"], f"{location}.pooled_record_count"
    )
    unweighted = _finite_number(
        counts["implicate_adjusted_unweighted_n"],
        f"{location}.implicate_adjusted_unweighted_n",
        minimum=0.0,
    )
    weighted = _finite_number(
        counts["weighted_business_interests"],
        f"{location}.weighted_business_interests",
        minimum=0.0,
    )
    if not math.isclose(
        unweighted,
        pooled / SCF_IMPLICATE_COUNT,
        rel_tol=0.0,
        abs_tol=1e-9,
    ):
        raise ValueError(f"{location} has inconsistent implicate-adjusted n.")
    if pooled == 0 and weighted != 0.0:
        raise ValueError(f"{location} gives weight to an empty sample.")
    if include_employer:
        employer_pooled = _nonnegative_integer(
            counts["employer_proxy_pooled_record_count"],
            f"{location}.employer_proxy_pooled_record_count",
        )
        employer_unweighted = _finite_number(
            counts["employer_proxy_implicate_adjusted_unweighted_n"],
            f"{location}.employer_proxy_implicate_adjusted_unweighted_n",
            minimum=0.0,
        )
        employer_weighted = _finite_number(
            counts["weighted_employer_proxy_business_interests"],
            f"{location}.weighted_employer_proxy_business_interests",
            minimum=0.0,
        )
        if employer_pooled > pooled or employer_weighted > weighted + 1e-9:
            raise ValueError(f"{location} employer counts exceed total counts.")
        if not math.isclose(
            employer_unweighted,
            employer_pooled / SCF_IMPLICATE_COUNT,
            rel_tol=0.0,
            abs_tol=1e-9,
        ):
            raise ValueError(
                f"{location} has inconsistent employer implicate-adjusted n."
            )
    return counts


def validate_qbi_employer_structure_resource(
    payload: Mapping[str, object],
) -> None:
    """Validate the committed SCF evidence-resource contract."""

    root = _exact_mapping(
        payload,
        {
            "schema_version",
            "resource_id",
            "provisional",
            "source",
            "methodology",
            "dimensions",
            "external_anchor",
            "cells",
            "profit_margin_quantiles",
            "provenance",
            "judgment_calls",
        },
        "QBI employer resource",
    )
    if root["schema_version"] != 1:
        raise ValueError("QBI employer resource schema_version must equal 1.")
    if root["resource_id"] != "qbi_employer_structure_v1":
        raise ValueError("QBI employer resource_id is not recognized.")
    if root["provisional"] is not True:
        raise ValueError("QBI employer resource must remain provisional.")

    source = _exact_mapping(
        root["source"],
        {
            "survey",
            "survey_year",
            "business_flow_year",
            "record_scope",
            "record_selection",
            "variables",
            "codebook",
        },
        "source",
    )
    if source["survey_year"] != 2022 or source["business_flow_year"] != 2021:
        raise ValueError("Employer source years must identify SCF 2022/flows 2021.")
    _nonempty_text(source["survey"], "source.survey")
    _nonempty_text(source["record_scope"], "source.record_scope")
    record_selection = _nonempty_text(
        source["record_selection"], "source.record_selection"
    )
    for variable in ("X3103", "X3104", "X3105"):
        if variable not in record_selection:
            raise ValueError(f"source.record_selection must name {variable}.")
    variables = _exact_mapping(
        source["variables"],
        {
            "active_management_screeners",
            "business_count",
            "weight",
            "industry",
            "headcount",
            "legal_form",
            "ownership_percent_x100",
            "gross_receipts",
            "whole_business_net_income",
            "respondent_or_spouse_works",
        },
        "source.variables",
    )
    if variables["active_management_screeners"] != ["X3103", "X3104"]:
        raise ValueError("source.variables has the wrong active-management screeners.")
    _exact_mapping(
        source["codebook"],
        {
            "url",
            "implicate_convention_lines",
            "weight_definition_lines",
            "business_variable_lines",
        },
        "source.codebook",
    )

    methodology = _exact_mapping(
        root["methodology"],
        {
            "point_estimate_weight",
            "implicate_count",
            "implicate_note",
            "income_measure",
            "employer_presence_proxy",
            "employer_proxy_limitation",
            "minimum_implicate_adjusted_unweighted_n",
            "presence_collapse_order",
            "size_collapse_order",
            "size_collapse_note",
            "headcount_disclosure",
        },
        "methodology",
    )
    if methodology["implicate_count"] != SCF_IMPLICATE_COUNT:
        raise ValueError("Employer methodology implicate_count must equal 5.")
    minimum_n = _finite_number(
        methodology["minimum_implicate_adjusted_unweighted_n"],
        "methodology.minimum_implicate_adjusted_unweighted_n",
        minimum=0.0,
    )
    if minimum_n == 0.0:
        raise ValueError("Employer methodology minimum n must be positive.")
    if methodology["presence_collapse_order"] != [
        "income_form_industry",
        "income_form",
        "form",
        "all",
    ]:
        raise ValueError("Employer presence collapse order is not recognized.")
    if methodology["size_collapse_order"] != [
        "income_form_industry_among_headcount_gt_1",
        "income_form_among_headcount_gt_1",
        "form_among_headcount_gt_1",
        "all_among_headcount_gt_1",
    ]:
        raise ValueError("Employer size collapse order is not recognized.")

    dimensions = _exact_mapping(
        root["dimensions"],
        {
            "income_bands",
            "legal_form_groups",
            "industry_bins",
            "headcount_size_bands",
        },
        "dimensions",
    )
    income_dimensions = _list(dimensions["income_bands"], "dimensions.income_bands")
    if [item.get("id") for item in income_dimensions if isinstance(item, Mapping)] != (
        list(SCF_INCOME_BANDS)
    ):
        raise ValueError("Employer income-band dimensions do not match the contract.")
    form_dimensions = _list(
        dimensions["legal_form_groups"], "dimensions.legal_form_groups"
    )
    if [item.get("id") for item in form_dimensions if isinstance(item, Mapping)] != (
        list(SCF_LEGAL_FORM_GROUPS)
    ):
        raise ValueError("Employer legal-form dimensions do not match the contract.")
    industry_dimensions = _list(dimensions["industry_bins"], "dimensions.industry_bins")
    if [
        item.get("code") for item in industry_dimensions if isinstance(item, Mapping)
    ] != list(SCF_INDUSTRY_BINS):
        raise ValueError("Employer industry dimensions do not match the contract.")
    size_dimensions = _list(
        dimensions["headcount_size_bands"], "dimensions.headcount_size_bands"
    )
    if [item.get("id") for item in size_dimensions if isinstance(item, Mapping)] != (
        list(SCF_HEADCOUNT_SIZE_BANDS)
    ):
        raise ValueError("Employer size-band dimensions do not match the contract.")
    _validate_provenance(root["provenance"], "provenance")
    _validate_text_list(root["judgment_calls"], "judgment_calls")

    fallback_levels = {"exact", "income_form", "form", "all"}

    def expected_dimensions(
        level: object,
        income_band: object,
        legal_form_group: object,
        industry_code: object,
    ) -> dict[str, object]:
        if level == "exact":
            return {
                "income_band": income_band,
                "legal_form_group": legal_form_group,
                "industry_code": industry_code,
            }
        if level == "income_form":
            return {
                "income_band": income_band,
                "legal_form_group": legal_form_group,
                "industry_code": "all",
            }
        if level == "form":
            return {
                "income_band": "all",
                "legal_form_group": legal_form_group,
                "industry_code": "all",
            }
        return {
            "income_band": "all",
            "legal_form_group": "all",
            "industry_code": "all",
        }

    cells = _list(root["cells"], "cells")
    expected_cell_count = (
        len(SCF_INCOME_BANDS) * len(SCF_LEGAL_FORM_GROUPS) * len(SCF_INDUSTRY_BINS)
    )
    if len(cells) != expected_cell_count:
        raise ValueError(
            f"Employer resource must carry {expected_cell_count} cells; "
            f"got {len(cells)}."
        )
    seen: set[tuple[object, object, object]] = set()
    for index, raw_cell in enumerate(cells):
        location = f"cells[{index}]"
        cell = _exact_mapping(
            raw_cell,
            {
                "income_band",
                "legal_form_group",
                "industry_code",
                "requested_counts",
                "employer_presence",
                "headcount_size_distribution",
            },
            location,
        )
        key = (
            cell["income_band"],
            cell["legal_form_group"],
            cell["industry_code"],
        )
        if key in seen:
            raise ValueError(f"Employer resource duplicates cell {key}.")
        seen.add(key)
        if key[0] not in SCF_INCOME_BANDS:
            raise ValueError(f"Employer cell has unknown income band {key[0]!r}.")
        if key[1] not in SCF_LEGAL_FORM_GROUPS:
            raise ValueError(f"Employer cell has unknown legal form {key[1]!r}.")
        if key[2] not in SCF_INDUSTRY_BINS:
            raise ValueError(f"Employer cell has unknown industry {key[2]!r}.")
        _validate_scf_counts(
            cell["requested_counts"],
            f"{location}.requested_counts",
            include_employer=True,
        )
        presence = _exact_mapping(
            cell["employer_presence"],
            {
                "probability_headcount_gt_1",
                "estimate_level",
                "source_dimensions",
                "source_counts",
            },
            f"{location}.employer_presence",
        )
        presence_level = presence["estimate_level"]
        if presence_level not in fallback_levels:
            raise ValueError(f"Employer cell {key} has unknown presence fallback.")
        if dict(
            _mapping(
                presence["source_dimensions"],
                f"{location}.employer_presence.source_dimensions",
            )
        ) != expected_dimensions(presence_level, *key):
            raise ValueError(f"Employer cell {key} has wrong presence dimensions.")
        presence_counts = _validate_scf_counts(
            presence["source_counts"],
            f"{location}.employer_presence.source_counts",
            include_employer=True,
        )
        if (
            presence_level != "all"
            and float(presence_counts["implicate_adjusted_unweighted_n"]) < minimum_n
        ):
            raise ValueError(f"Employer cell {key} retains a thin presence sample.")
        denominator = float(presence_counts["weighted_business_interests"])
        if denominator <= 0.0:
            raise ValueError(f"Employer cell {key} has zero presence denominator.")
        probability = _finite_probability(
            presence["probability_headcount_gt_1"],
            f"{location}.employer_presence.probability",
        )
        expected_probability = (
            float(presence_counts["weighted_employer_proxy_business_interests"])
            / denominator
        )
        if not math.isclose(
            probability, expected_probability, rel_tol=0.0, abs_tol=1e-12
        ):
            raise ValueError(f"Employer cell {key} probability disagrees with counts.")

        size = _exact_mapping(
            cell["headcount_size_distribution"],
            {
                "conditional_on",
                "shares",
                "estimate_level",
                "source_dimensions",
                "source_counts",
            },
            f"{location}.headcount_size_distribution",
        )
        if size["conditional_on"] != "headcount_gt_1":
            raise ValueError(f"Employer cell {key} has wrong size conditioning.")
        size_level = size["estimate_level"]
        if size_level not in fallback_levels:
            raise ValueError(f"Employer cell {key} has unknown size fallback.")
        if dict(
            _mapping(
                size["source_dimensions"],
                f"{location}.headcount_size_distribution.source_dimensions",
            )
        ) != expected_dimensions(size_level, *key):
            raise ValueError(f"Employer cell {key} has wrong size dimensions.")
        size_counts = _validate_scf_counts(
            size["source_counts"],
            f"{location}.headcount_size_distribution.source_counts",
            include_employer=True,
        )
        if size_level != "all" and (
            float(size_counts["implicate_adjusted_unweighted_n"]) < minimum_n
        ):
            raise ValueError(f"Employer cell {key} retains a thin size sample.")
        if size_counts["pooled_record_count"] != size_counts[
            "employer_proxy_pooled_record_count"
        ] or not math.isclose(
            float(size_counts["weighted_business_interests"]),
            float(size_counts["weighted_employer_proxy_business_interests"]),
            rel_tol=0.0,
            abs_tol=1e-9,
        ):
            raise ValueError(f"Employer cell {key} size sample includes nonemployers.")
        shares = _exact_mapping(
            size["shares"],
            set(SCF_HEADCOUNT_SIZE_BANDS),
            f"{location}.headcount_size_distribution.shares",
        )
        values = [
            _finite_probability(shares[band], f"{location}.shares.{band}")
            for band in SCF_HEADCOUNT_SIZE_BANDS
        ]
        if not math.isclose(sum(values), 1.0, rel_tol=0.0, abs_tol=1e-9):
            raise ValueError(f"Employer cell {key} size shares do not sum to 1.")

    external = _exact_mapping(
        root["external_anchor"],
        {
            "source",
            "source_url",
            "tax_year",
            "zero_employee_firm_share",
            "zero_employee_deduction_dollar_share",
            "sole_proprietor_zero_employee_share_statement",
            "partnership_zero_employee_share_statement",
            "scf_comparison",
            "definition_gaps",
        },
        "external_anchor",
    )
    if external["tax_year"] != 2022:
        raise ValueError("External JCT anchor must identify tax year 2022.")
    _finite_probability(
        external["zero_employee_firm_share"],
        "external_anchor.zero_employee_firm_share",
    )
    _finite_probability(
        external["zero_employee_deduction_dollar_share"],
        "external_anchor.zero_employee_deduction_dollar_share",
    )
    _validate_text_list(external["definition_gaps"], "external_anchor.definition_gaps")
    comparisons = _exact_mapping(
        external["scf_comparison"],
        {
            "main_proxy_including_informal_code_40",
            "strict_form_sensitivity_excluding_code_40",
        },
        "external_anchor.scf_comparison",
    )
    expected_comparison_codes = {
        "main_proxy_including_informal_code_40": sorted(_MAIN_QBI_PROXY_CODES),
        "strict_form_sensitivity_excluding_code_40": sorted(_STRICT_QBI_PROXY_CODES),
    }
    for name, raw_comparison in comparisons.items():
        location = f"external_anchor.scf_comparison.{name}"
        comparison = _exact_mapping(
            raw_comparison,
            {
                "sample_definition",
                "numerator_definition",
                "legal_form_codes",
                "pooled_record_count",
                "implicate_adjusted_unweighted_n",
                "weighted_business_interests",
                "headcount_le_1_share",
                "difference_from_jct_percentage_points",
            },
            location,
        )
        _nonempty_text(comparison["sample_definition"], f"{location}.sample_definition")
        _nonempty_text(
            comparison["numerator_definition"], f"{location}.numerator_definition"
        )
        if comparison["legal_form_codes"] != expected_comparison_codes[name]:
            raise ValueError(f"{location} has the wrong legal-form sample.")
        _validate_scf_counts(
            {key: comparison[key] for key in _SCF_SIMPLE_COUNT_KEYS},
            f"{location}.counts",
            include_employer=False,
        )
        share = _finite_probability(
            comparison["headcount_le_1_share"],
            f"{location}.headcount_le_1_share",
        )
        difference = _finite_number(
            comparison["difference_from_jct_percentage_points"],
            f"{location}.difference_from_jct_percentage_points",
        )
        if not math.isclose(
            difference,
            100.0 * (share - JCT_ZERO_EMPLOYEE_FIRM_SHARE),
            rel_tol=0.0,
            abs_tol=1e-12,
        ):
            raise ValueError(f"{location} has an inconsistent JCT difference.")

    margins = _exact_mapping(
        root["profit_margin_quantiles"],
        {
            "definition",
            "ownership_note",
            "support_note",
            "quantile_method",
            "probabilities",
            "minimum_implicate_adjusted_unweighted_n",
            "collapse_order",
            "cells",
        },
        "profit_margin_quantiles",
    )
    if margins["probabilities"] != list(SCF_MARGIN_QUANTILES):
        raise ValueError("Profit-margin probabilities do not match the contract.")
    if margins["minimum_implicate_adjusted_unweighted_n"] != minimum_n:
        raise ValueError("Profit-margin and employer minimum n values differ.")
    if margins["collapse_order"] != ["form_industry", "form", "all"]:
        raise ValueError("Profit-margin collapse order is not recognized.")
    margin_cells = _list(margins["cells"], "profit_margin_quantiles.cells")
    expected_margin_count = len(SCF_LEGAL_FORM_GROUPS) * len(SCF_INDUSTRY_BINS)
    if len(margin_cells) != expected_margin_count:
        raise ValueError(
            f"Profit-margin resource must carry {expected_margin_count} cells."
        )
    margin_seen: set[tuple[object, object]] = set()
    quantile_names = [_quantile_name(q) for q in SCF_MARGIN_QUANTILES]
    for index, raw_cell in enumerate(margin_cells):
        location = f"profit_margin_quantiles.cells[{index}]"
        cell = _exact_mapping(
            raw_cell,
            {
                "legal_form_group",
                "industry_code",
                "requested_counts",
                "estimate_level",
                "source_dimensions",
                "source_counts",
                "quantiles",
                "source_minimum",
                "source_maximum",
            },
            location,
        )
        key = (cell["legal_form_group"], cell["industry_code"])
        if key in margin_seen:
            raise ValueError(f"Profit-margin resource duplicates cell {key}.")
        margin_seen.add(key)
        if key[0] not in SCF_LEGAL_FORM_GROUPS or key[1] not in SCF_INDUSTRY_BINS:
            raise ValueError(f"Profit-margin resource has unknown cell {key}.")
        _validate_scf_counts(
            cell["requested_counts"],
            f"{location}.requested_counts",
            include_employer=False,
        )
        level = cell["estimate_level"]
        if level not in {"exact", "form", "all"}:
            raise ValueError(f"Profit-margin cell {key} has unknown fallback.")
        expected_margin_dimensions = {
            "legal_form_group": key[0] if level != "all" else "all",
            "industry_code": key[1] if level == "exact" else "all",
        }
        if (
            dict(_mapping(cell["source_dimensions"], f"{location}.source_dimensions"))
            != expected_margin_dimensions
        ):
            raise ValueError(f"Profit-margin cell {key} has wrong source dimensions.")
        source_counts = _validate_scf_counts(
            cell["source_counts"],
            f"{location}.source_counts",
            include_employer=False,
        )
        if (
            level != "all"
            and float(source_counts["implicate_adjusted_unweighted_n"]) < minimum_n
        ):
            raise ValueError(f"Profit-margin cell {key} retains a thin sample.")
        if float(source_counts["weighted_business_interests"]) <= 0.0:
            raise ValueError(f"Profit-margin cell {key} has zero source weight.")
        quantiles = _exact_mapping(
            cell["quantiles"], set(quantile_names), f"{location}.quantiles"
        )
        if list(quantiles) != quantile_names:
            raise ValueError(
                f"Profit-margin cell {key} quantiles must be ordered {quantile_names}."
            )
        values: list[float] = []
        for name in quantile_names:
            values.append(_finite_number(quantiles[name], f"{location}.{name}"))
        if values != sorted(values):
            raise ValueError(f"Profit-margin cell {key} quantiles are not monotone.")
        source_minimum = _finite_number(
            cell["source_minimum"], f"{location}.source_minimum"
        )
        source_maximum = _finite_number(
            cell["source_maximum"], f"{location}.source_maximum"
        )
        if source_minimum > values[0] or values[-1] > source_maximum:
            raise ValueError(f"Profit-margin cell {key} quantiles exceed source range.")


@dataclass(frozen=True)
class SoiIndustryObservation:
    """One published SOI industry column or row before ratio derivation."""

    form: str
    tax_year: int
    published_label: str
    industry_path: tuple[str, ...]
    source_ordinal: int
    industry_level: str
    is_aggregate: bool
    receipts: float | None
    salaries: float | None
    cost_labor: float | None
    officer_compensation: float | None
    guaranteed_payments_excluded: float | None
    payroll: float | None
    gross_depreciable_assets: float | None
    depreciation_deduction: float | None
    publication_flags: Mapping[str, str]
    provenance: Mapping[str, object]

    def __post_init__(self) -> None:
        if self.form not in SOI_ENTITY_FORMS:
            raise ValueError(f"Unknown SOI entity form {self.form!r}.")
        if self.tax_year not in (2022, 2023):
            raise ValueError(f"Unsupported SOI tax year {self.tax_year!r}.")
        if not self.published_label.strip() or not self.industry_path:
            raise ValueError("SOI observations require a published industry path.")
        if self.source_ordinal <= 0:
            raise ValueError("SOI source ordinals must be positive.")
        if self.industry_level not in {
            "all",
            "sector_total",
            "published_detail",
            "unallocable",
        }:
            raise ValueError(f"Unknown SOI industry level {self.industry_level!r}.")
        unknown_flags = set(self.publication_flags.values()) - SOI_PUBLICATION_FLAGS
        if unknown_flags:
            raise ValueError(
                f"SOI observation has unknown publication flags "
                f"{sorted(unknown_flags)}."
            )
        for name in (
            "receipts",
            "salaries",
            "cost_labor",
            "officer_compensation",
            "guaranteed_payments_excluded",
            "payroll",
            "gross_depreciable_assets",
            "depreciation_deduction",
        ):
            value = getattr(self, name)
            if value is not None and not math.isfinite(value):
                raise ValueError(f"SOI observation {name} must be finite or null.")


def _clean_label(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\u2014", "-").split())


def _comparable_label(value: object) -> str:
    return re.sub(r"\s*\[\d+\]\s*$", "", _clean_label(value))


def _excel_column(column: int) -> str:
    if column <= 0:
        raise ValueError("Excel columns are one-based positive integers.")
    letters = ""
    while column:
        column, remainder = divmod(column - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def _cell_reference(sheet: str, row: int, column: int) -> str:
    return f"{sheet}!{_excel_column(column)}{row}"


def _format_flag(format_string: str, value: object) -> str:
    text = str(value).strip().lower() if isinstance(value, str) else ""
    if text in {"d", "[d]"}:
        return "deleted_for_disclosure"
    if text.startswith("**") or '"** "' in format_string:
        return "combined_for_disclosure"
    if text.startswith("*") or '"* "' in format_string:
        return "caution"
    return "published"


def _numeric_with_flag(
    value: object,
    flag: str,
) -> float | None:
    if flag in {"combined_for_disclosure", "deleted_for_disclosure"}:
        return None
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        return None
    numeric = float(value)
    return numeric if math.isfinite(numeric) else None


def _combined_flag(flags: Sequence[str]) -> str:
    if "deleted_for_disclosure" in flags:
        return "deleted_for_disclosure"
    if "combined_for_disclosure" in flags:
        return "combined_for_disclosure"
    if "not_published" in flags:
        return "not_published"
    if "caution" in flags:
        return "caution"
    return "published"


def _ratio(
    observation: SoiIndustryObservation,
    *,
    numerator_names: Sequence[str],
) -> tuple[float | None, str, str | None]:
    receipt_flag = observation.publication_flags.get("receipts", "not_published")
    flags = [receipt_flag]
    numerator = 0.0
    for name in numerator_names:
        flags.append(observation.publication_flags.get(name, "not_published"))
        value = getattr(observation, name)
        if value is None:
            return None, _combined_flag(flags), f"{name}_not_published"
        numerator += value
    if observation.receipts is None:
        return None, _combined_flag(flags), "receipts_not_published"
    if observation.receipts <= 0.0:
        return None, _combined_flag(flags), "nonpositive_receipts"
    return (
        float(numerator / observation.receipts),
        _combined_flag(flags),
        None,
    )


def census_bin_hint(
    industry_path: Sequence[str],
) -> tuple[int | None, str]:
    """Map a published SOI label conservatively to an SCF seven-bin hint."""

    path = tuple(_clean_label(label) for label in industry_path if _clean_label(label))
    if not path:
        return None, "No published industry label."
    label = path[-1].lower()
    text = " > ".join(path).lower()
    if any(
        token in text
        for token in (
            "all industries",
            "all nonfarm industries",
            "unclassified",
            "not allocable",
            "unallocable",
        )
    ):
        return None, "Aggregate or unclassified category has no SCF-bin hint."

    if "veterinary" in label or "landscap" in label:
        return 1, "Exact SCF bin-1 specialization."
    if "agriculture" in text or "forestry" in text or "fishing" in text:
        return 1, "SCF bin 1 covers the published agriculture group."
    if "mining" in text:
        return 2, "SCF bin 2 covers mining."
    if "construction" in text:
        return 2, "SCF bin 2 covers construction."
    if "manufactur" in text:
        return 3, "SCF bin 3 covers manufacturing."
    if "wholesale" in text or "retail" in text:
        return 4, "SCF bin 4 covers wholesale and retail trade."

    if "information" in text or "publishing" in text:
        if "software publishing" in label:
            return 5, "Software publishing is an exact SCF bin-5 specialization."
        if any(
            token in label for token in ("newspaper", "periodical", "book", "directory")
        ):
            return 3, "Selected publishing is an exact SCF bin-3 specialization."
        if "data processing" in label or "hosting" in label:
            return 5, "Data processing and hosting map to SCF bin 5."
        if label in {"information", "publishing industries"} or label == "total":
            return None, "Published information aggregate spans SCF bins 3, 5, and 6."
        return 6, "Remaining detailed information services map to SCF bin 6."

    if "finance" in text or "insurance" in text:
        return 5, "SCF bin 5 covers finance and insurance."
    if "real estate" in label and ("rental" in label or "leasing" in label):
        return None, "Combined real-estate and rental total spans SCF bins 5 and 6."
    if "real estate" in label:
        return 5, "SCF bin 5 covers real estate."
    if "rental" in text or "leasing" in text:
        if any(
            token in label
            for token in (
                "automotive",
                "commercial",
                "industrial",
                "intangible",
                "lessors of nonfinancial",
            )
        ):
            return 5, "Exact rental specialization maps to SCF bin 5."
        if any(
            token in label
            for token in (
                "consumer",
                "formal wear",
                "video",
                "home health",
                "recreational",
                "general rental",
            )
        ):
            return 6, "Exact consumer-rental specialization maps to SCF bin 6."
        return None, "Published rental aggregate spans SCF bins 5 and 6."

    if "utility" in text or "utilities" in text:
        return 6, "SCF bin 6 covers utilities."
    if "transportation" in text or "warehousing" in text:
        return 6, "SCF bin 6 covers transportation and warehousing."
    if "professional" in text or "scientific" in text or "technical" in text:
        if label in {
            "professional, scientific, and technical services",
            "professional services",
            "other professional, scientific, and technical services",
            "other miscellaneous services",
            "total",
        }:
            return None, "Broad professional-services total can include veterinary."
        return 6, "Detailed non-veterinary professional services map to SCF bin 6."
    if "management of companies" in text or "holding companies" in text:
        return 6, "SCF bin 6 covers management of companies."

    if "administrative" in text or "support" in text or "waste" in text:
        if "landscap" in label:
            return 1, "Landscaping is an exact SCF bin-1 specialization."
        if any(
            token in label
            for token in (
                "employment",
                "business support",
                "investigation",
                "security",
            )
        ):
            return 5, "Exact administrative specialization maps to SCF bin 5."
        if label in {
            "administrative and support and waste management and remediation services",
            "administrative and support services",
            "total",
        }:
            return None, "Broad administrative total spans SCF bins 1, 5, and 6."
        return 6, "Remaining detailed administrative or waste services map to bin 6."

    if "accommodation" in label and "food service" in label:
        return None, "Combined accommodation and food total spans bins 4 and 6."
    if "restaurant" in label or "drinking place" in label or "food service" in label:
        return 4, "Restaurants and drinking places map to SCF bin 4."
    if "accommodation" in text or any(
        token in label for token in ("hotel", "motel", "rooming", "boarding")
    ):
        return 6, "Accommodation maps to SCF bin 6."

    if "repair" in text or "maintenance" in text:
        return 5, "Repair and maintenance map to SCF bin 5."
    if "educational" in label and "other services" in label:
        return None, "Combined educational and other-services total spans bins 5 and 6."
    if "other services" in text:
        if label in {"other services", "total"}:
            return None, "Other-services total spans SCF bins 5 and 6."
        return 6, "Detailed non-repair other services map to SCF bin 6."

    if "arts" in text or "entertainment" in text or "recreation" in text:
        if "museum" in label or "amusement" in label:
            return 6, "Museums and amusement map to SCF bin 6."
        return None, "The SCF codebook's exact 8560 arts seam is ambiguous."
    if any(
        token in text
        for token in (
            "education",
            "health care",
            "social assistance",
            "personal and laundry",
            "religious",
            "private household",
        )
    ):
        return 6, "Published service group maps to SCF bin 6."
    if "public administration" in text:
        return 7, "SCF bin 7 covers public administration."
    return None, "Published label cannot be mapped deterministically to an SCF bin."


def _xlrd_format_string(book: object, sheet: object, row: int, column: int) -> str:
    xf_index = sheet.cell_xf_index(row, column)
    xf = book.xf_list[xf_index]
    return book.format_map[xf.format_key].format_str


def _xlrd_header_anchors(
    sheet: object,
    *,
    column: int,
    rows: range,
) -> list[str]:
    coordinates: list[str] = []
    for row in rows:
        anchor_row, anchor_column = row, column
        for row_low, row_high, column_low, column_high in sheet.merged_cells:
            if row_low <= row < row_high and column_low <= column < column_high:
                anchor_row, anchor_column = row_low, column_low
                break
        reference = _cell_reference(sheet.name, anchor_row + 1, anchor_column + 1)
        label = _clean_label(sheet.cell_value(anchor_row, anchor_column))
        if label and reference not in coordinates:
            coordinates.append(reference)
    return coordinates


def parse_sole_proprietor_soi_workbooks(
    business_table_path: Path | str,
    income_statement_path: Path | str,
) -> list[SoiIndustryObservation]:
    """Parse SOI sole-proprietor Tables 1 and 2 with disclosure flags."""

    try:
        import xlrd
    except ImportError as error:  # pragma: no cover - dependency contract
        raise RuntimeError(
            "Reading IRS SOI .xls workbooks requires the declared xlrd dependency."
        ) from error

    business_path = Path(business_table_path)
    income_path = Path(income_statement_path)
    business_book = xlrd.open_workbook(business_path, formatting_info=True)
    income_book = xlrd.open_workbook(income_path, formatting_info=True)
    if business_book.sheet_names() != ["TAB1"]:
        raise ValueError("Sole-proprietor business workbook must contain TAB1.")
    if income_book.sheet_names() != ["TAB2"]:
        raise ValueError("Sole-proprietor income workbook must contain TAB2.")
    tab1 = business_book.sheet_by_name("TAB1")
    tab2 = income_book.sheet_by_name("TAB2")
    if "Tax Year 2023" not in str(tab1.cell_value(0, 0)):
        raise ValueError("Sole-proprietor TAB1 title does not identify tax year 2023.")
    if "Tax Year 2023" not in str(tab2.cell_value(0, 0)):
        raise ValueError("Sole-proprietor TAB2 title does not identify tax year 2023.")

    start_row = next(
        row
        for row in range(tab1.nrows)
        if _clean_label(tab1.cell_value(row, 0)) == "All nonfarm industries"
    )
    end_row = next(
        row
        for row in range(start_row + 1, tab1.nrows)
        if _clean_label(tab1.cell_value(row, 0)).startswith(
            "Estimate should be used with caution"
        )
    )
    source_rows = list(range(start_row, end_row))
    tab2_columns: list[int] = []
    expected = 1
    for column in range(1, tab2.ncols):
        value = tab2.cell_value(9, column)
        if isinstance(value, (int, float)) and int(value) == expected:
            tab2_columns.append(column)
            expected += 1
        elif tab2_columns:
            break
    if len(source_rows) != len(tab2_columns):
        raise ValueError(
            "Sole-proprietor table alignment failed: TAB1 industry rows and "
            "TAB2 industry columns differ."
        )

    paths: list[tuple[str, ...]] = []
    indents: list[int] = []
    stack: list[tuple[int, str]] = []
    for position, row in enumerate(source_rows):
        raw_label = str(tab1.cell_value(row, 0))
        label = _clean_label(raw_label)
        indent = len(raw_label) - len(raw_label.lstrip())
        if position == 0:
            path = (label,)
            stack = []
        else:
            while stack and stack[-1][0] >= indent:
                stack.pop()
            path = tuple(item[1] for item in stack) + (label,)
            stack.append((indent, label))
        paths.append(path)
        indents.append(indent)

    observations: list[SoiIndustryObservation] = []
    for position, (row, tab2_column, path) in enumerate(
        zip(source_rows, tab2_columns, paths, strict=True)
    ):
        is_all = position == 0
        has_child = (
            position + 1 < len(indents) and indents[position + 1] > indents[position]
        )
        is_unclassified = path[-1].lower() == "unclassified establishments"
        is_aggregate = is_all or has_child
        level = (
            "all"
            if is_all
            else (
                "sector_total"
                if has_child
                else ("unallocable" if is_unclassified else "published_detail")
            )
        )

        tab1_cells = {
            "receipts": (row, 2),
            "depreciation_deduction": (row, 3),
            "payroll": (row, 7),
        }
        tab2_cells = {
            "receipts_crosscheck": (12, tab2_column),
            "cost_labor": (18, tab2_column),
            "salaries": (41, tab2_column),
        }
        values: dict[str, float | None] = {}
        flags: dict[str, str] = {}
        for name, (cell_row, cell_column) in tab1_cells.items():
            raw = tab1.cell_value(cell_row, cell_column)
            flag = _format_flag(
                _xlrd_format_string(business_book, tab1, cell_row, cell_column),
                raw,
            )
            values[name] = _numeric_with_flag(raw, flag)
            flags[name] = flag
        component_values: dict[str, float | None] = {}
        for name, (cell_row, cell_column) in tab2_cells.items():
            raw = tab2.cell_value(cell_row, cell_column)
            flag = _format_flag(
                _xlrd_format_string(income_book, tab2, cell_row, cell_column),
                raw,
            )
            component_values[name] = _numeric_with_flag(raw, flag)
            flags[name] = flag

        receipt_raw = tab1.cell_value(row, 2)
        tab2_receipt_raw = tab2.cell_value(12, tab2_column)
        if not math.isclose(
            float(receipt_raw),
            float(tab2_receipt_raw),
            rel_tol=0.0,
            abs_tol=0.0,
        ):
            raise ValueError(
                f"Sole-proprietor receipts mismatch at source ordinal {position + 1}."
            )
        labor_raw = tab2.cell_value(18, tab2_column)
        salary_raw = tab2.cell_value(41, tab2_column)
        payroll_raw = tab1.cell_value(row, 7)
        if all(
            isinstance(value, (int, float))
            for value in (labor_raw, salary_raw, payroll_raw)
        ) and not math.isclose(
            float(labor_raw) + float(salary_raw),
            float(payroll_raw),
            rel_tol=0.0,
            abs_tol=0.5,
        ):
            raise ValueError(
                f"Sole-proprietor payroll identity failed at source ordinal "
                f"{position + 1}."
            )

        observations.append(
            SoiIndustryObservation(
                form="sole_proprietorship",
                tax_year=2023,
                published_label=path[-1],
                industry_path=path,
                source_ordinal=position + 1,
                industry_level=level,
                is_aggregate=is_aggregate,
                receipts=values["receipts"],
                salaries=component_values["salaries"],
                cost_labor=component_values["cost_labor"],
                officer_compensation=None,
                guaranteed_payments_excluded=None,
                payroll=values["payroll"],
                gross_depreciable_assets=None,
                depreciation_deduction=values["depreciation_deduction"],
                publication_flags={
                    "receipts": flags["receipts"],
                    "salaries": flags["salaries"],
                    "cost_labor": flags["cost_labor"],
                    "officer_compensation": "not_published",
                    "guaranteed_payments_excluded": "not_published",
                    "payroll": flags["payroll"],
                    "gross_depreciable_assets": "not_published",
                    "depreciation_deduction": flags["depreciation_deduction"],
                },
                provenance={
                    "source_tables": [
                        f"{business_path.name} Table 1",
                        f"{income_path.name} Table 2",
                    ],
                    "sheet_names": ["TAB1", "TAB2"],
                    "tax_year": 2023,
                    "units": "thousands_of_dollars",
                    "industry_cells": [
                        _cell_reference("TAB1", row + 1, 1),
                        *_xlrd_header_anchors(
                            tab2, column=tab2_column, rows=range(2, 9)
                        ),
                    ],
                    "receipts_cell": _cell_reference("TAB1", row + 1, 3),
                    "wage_cells": [_cell_reference("TAB1", row + 1, 8)],
                    "wage_component_crosscheck_cells": [
                        _cell_reference("TAB2", 19, tab2_column + 1),
                        _cell_reference("TAB2", 42, tab2_column + 1),
                    ],
                    "capital_cell": _cell_reference("TAB1", row + 1, 4),
                    "calculation": {
                        "wage_share": "TAB1 payroll / TAB1 business receipts",
                        "ubia_intensity": (
                            "TAB1 depreciation deduction / TAB1 business "
                            "receipts (flow proxy)"
                        ),
                    },
                },
            )
        )
    return observations


def _xlsx_anchor(sheet: object, row: int, column: int) -> tuple[int, int]:
    for merged_range in sheet.merged_cells.ranges:
        if (
            merged_range.min_row <= row <= merged_range.max_row
            and merged_range.min_col <= column <= merged_range.max_col
        ):
            return merged_range.min_row, merged_range.min_col
    return row, column


def _xlsx_path(
    sheet: object,
    *,
    column: int,
    rows: range,
) -> tuple[tuple[str, ...], list[str]]:
    labels: list[str] = []
    references: list[str] = []
    for row in rows:
        anchor_row, anchor_column = _xlsx_anchor(sheet, row, column)
        label = _clean_label(sheet.cell(anchor_row, anchor_column).value)
        reference = _cell_reference(sheet.title, anchor_row, anchor_column)
        if label and (not labels or labels[-1] != label):
            labels.append(label)
        if label and reference not in references:
            references.append(reference)
    return tuple(labels), references


def _xlsx_amount(cell: object) -> tuple[float | None, str]:
    flag = _format_flag(str(cell.number_format), cell.value)
    return _numeric_with_flag(cell.value, flag), flag


def parse_partnership_soi_workbooks(
    income_table_path: Path | str,
    balance_sheet_path: Path | str,
) -> list[SoiIndustryObservation]:
    """Parse sector-level partnership wage and depreciable-asset anchors."""

    from openpyxl import load_workbook

    income_path = Path(income_table_path)
    balance_path = Path(balance_sheet_path)
    income_book = load_workbook(income_path, data_only=True, read_only=False)
    balance_book = load_workbook(balance_path, data_only=True, read_only=False)
    if income_book.sheetnames != ["Sheet1"] or balance_book.sheetnames != ["Sheet1"]:
        raise ValueError("Partnership workbooks must contain exactly Sheet1.")
    income = income_book["Sheet1"]
    balance = balance_book["Sheet1"]
    if "Tax Year 2023" not in str(income["A1"].value):
        raise ValueError("Partnership income title does not identify tax year 2023.")
    if "Tax Year 2023" not in str(balance["A1"].value):
        raise ValueError("Partnership balance title does not identify tax year 2023.")

    observations: list[SoiIndustryObservation] = []
    for column in range(2, 22):
        label = _clean_label(income.cell(4, column).value)
        balance_label = _clean_label(balance.cell(4, column).value)
        if _comparable_label(label) != _comparable_label(balance_label):
            raise ValueError(
                f"Partnership workbook industry mismatch in column "
                f"{_excel_column(column)}."
            )
        is_all = column == 2
        is_unallocable = column == 21
        level = (
            "all" if is_all else ("unallocable" if is_unallocable else "sector_total")
        )
        cells = {
            "receipts": income.cell(18, column),
            "cost_labor": income.cell(26, column),
            "salaries": income.cell(30, column),
            "guaranteed_payments_excluded": income.cell(31, column),
            "depreciation_deduction": income.cell(37, column),
            "gross_depreciable_assets": balance.cell(29, column),
        }
        values: dict[str, float | None] = {}
        flags: dict[str, str] = {}
        for name, cell in cells.items():
            values[name], flags[name] = _xlsx_amount(cell)
        observations.append(
            SoiIndustryObservation(
                form="partnership",
                tax_year=2023,
                published_label=label,
                industry_path=(label,),
                source_ordinal=column - 1,
                industry_level=level,
                is_aggregate=is_all,
                receipts=values["receipts"],
                salaries=values["salaries"],
                cost_labor=values["cost_labor"],
                officer_compensation=None,
                guaranteed_payments_excluded=values["guaranteed_payments_excluded"],
                payroll=None,
                gross_depreciable_assets=values["gross_depreciable_assets"],
                depreciation_deduction=values["depreciation_deduction"],
                publication_flags={
                    "receipts": flags["receipts"],
                    "salaries": flags["salaries"],
                    "cost_labor": flags["cost_labor"],
                    "officer_compensation": "not_published",
                    "guaranteed_payments_excluded": flags[
                        "guaranteed_payments_excluded"
                    ],
                    "payroll": "not_published",
                    "gross_depreciable_assets": flags["gross_depreciable_assets"],
                    "depreciation_deduction": flags["depreciation_deduction"],
                },
                provenance={
                    "source_tables": [
                        f"{income_path.name} Table 1",
                        f"{balance_path.name} Table 3",
                    ],
                    "sheet_names": ["Sheet1", "Sheet1"],
                    "tax_year": 2023,
                    "units": "thousands_of_dollars",
                    "industry_cells": [
                        f"{income_path.name}:{_cell_reference('Sheet1', 4, column)}",
                        f"{balance_path.name}:{_cell_reference('Sheet1', 4, column)}",
                    ],
                    "receipts_cell": (
                        f"{income_path.name}:{_cell_reference('Sheet1', 18, column)}"
                    ),
                    "wage_cells": [
                        f"{income_path.name}:{_cell_reference('Sheet1', 26, column)}",
                        f"{income_path.name}:{_cell_reference('Sheet1', 30, column)}",
                    ],
                    "excluded_wage_cell": (
                        f"{income_path.name}:{_cell_reference('Sheet1', 31, column)}"
                    ),
                    "capital_cell": (
                        f"{balance_path.name}:{_cell_reference('Sheet1', 29, column)}"
                    ),
                    "depreciation_deduction_cell": (
                        f"{income_path.name}:{_cell_reference('Sheet1', 37, column)}"
                    ),
                    "calculation": {
                        "wage_share": (
                            "(cost of labor + salaries and wages) / business receipts"
                        ),
                        "ubia_intensity": (
                            "gross depreciable assets / business receipts"
                        ),
                    },
                },
            )
        )
    return observations


def parse_s_corporation_soi_workbook(
    workbook_path: Path | str,
) -> list[SoiIndustryObservation]:
    """Parse Form 1120-S Table 6.1 at its published major-industry detail."""

    from openpyxl import load_workbook

    path = Path(workbook_path)
    book = load_workbook(path, data_only=True, read_only=False)
    if book.sheetnames != ["Table 6.1"]:
        raise ValueError("S-corporation workbook must contain Table 6.1.")
    sheet = book["Table 6.1"]
    if "Tax Year 2022" not in str(sheet["A2"].value):
        raise ValueError("S-corporation title does not identify tax year 2022.")

    columns: list[int] = []
    expected = 1
    for column in range(2, sheet.max_column + 1):
        value = sheet.cell(8, column).value
        if isinstance(value, (int, float)) and int(value) == expected:
            columns.append(column)
            expected += 1
        elif columns:
            break
    if not columns:
        raise ValueError("S-corporation workbook has no numbered industries.")

    observations: list[SoiIndustryObservation] = []
    for position, column in enumerate(columns):
        path_labels, header_cells = _xlsx_path(sheet, column=column, rows=range(5, 8))
        if not path_labels:
            raise ValueError(
                f"S-corporation column {_excel_column(column)} has no header."
            )
        is_all = position == 0
        is_total = any(label.lower() == "total" for label in path_labels)
        level = (
            "all" if is_all else ("sector_total" if is_total else "published_detail")
        )
        cells = {
            "gross_depreciable_assets": sheet.cell(21, column),
            "receipts": sheet.cell(43, column),
            "officer_compensation": sheet.cell(49, column),
            "salaries": sheet.cell(50, column),
            "depreciation_deduction": sheet.cell(57, column),
        }
        values: dict[str, float | None] = {}
        flags: dict[str, str] = {}
        for name, cell in cells.items():
            values[name], flags[name] = _xlsx_amount(cell)
        observations.append(
            SoiIndustryObservation(
                form="s_corporation",
                tax_year=2022,
                published_label=path_labels[-1],
                industry_path=path_labels,
                source_ordinal=position + 1,
                industry_level=level,
                is_aggregate=is_all or is_total,
                receipts=values["receipts"],
                salaries=values["salaries"],
                cost_labor=None,
                officer_compensation=values["officer_compensation"],
                guaranteed_payments_excluded=None,
                payroll=None,
                gross_depreciable_assets=values["gross_depreciable_assets"],
                depreciation_deduction=values["depreciation_deduction"],
                publication_flags={
                    "receipts": flags["receipts"],
                    "salaries": flags["salaries"],
                    "cost_labor": "not_published",
                    "officer_compensation": flags["officer_compensation"],
                    "guaranteed_payments_excluded": "not_published",
                    "payroll": "not_published",
                    "gross_depreciable_assets": flags["gross_depreciable_assets"],
                    "depreciation_deduction": flags["depreciation_deduction"],
                },
                provenance={
                    "source_tables": [f"{path.name} Table 6.1"],
                    "sheet_names": ["Table 6.1"],
                    "tax_year": 2022,
                    "units": "thousands_of_dollars",
                    "industry_cells": header_cells,
                    "receipts_cell": _cell_reference("Table 6.1", 43, column),
                    "wage_cells": [
                        _cell_reference("Table 6.1", 49, column),
                        _cell_reference("Table 6.1", 50, column),
                    ],
                    "capital_cell": _cell_reference("Table 6.1", 21, column),
                    "depreciation_deduction_cell": _cell_reference(
                        "Table 6.1", 57, column
                    ),
                    "calculation": {
                        "wage_share": (
                            "(compensation of officers + salaries and wages) / "
                            "business receipts"
                        ),
                        "ubia_intensity": (
                            "gross depreciable assets / business receipts"
                        ),
                    },
                },
            )
        )
    return observations


def inspect_all_corporation_soi_workbook(
    workbook_path: Path | str,
) -> dict[str, object]:
    """Validate and document why all-corporation Table 5.1 is not used."""

    from openpyxl import load_workbook

    path = Path(workbook_path)
    book = load_workbook(path, data_only=True, read_only=False)
    if book.sheetnames != ["Table 5.1"]:
        raise ValueError("All-corporation workbook must contain Table 5.1.")
    title = _clean_label(book["Table 5.1"]["A2"].value)
    if "Tax Year 2022" not in title or "Minor Industry" not in title:
        raise ValueError("All-corporation Table 5.1 title is not recognized.")
    return {
        "filename": path.name,
        "table": "Table 5.1",
        "tax_year": 2022,
        "review_status": "inspected_not_used",
        "reason": (
            "The table covers all active corporations, including C corporations. "
            "Its finer minor-industry detail cannot identify S-corporation priors; "
            "Form 1120-S Table 6.1 is used instead."
        ),
    }


def _industry_key(observation: SoiIndustryObservation) -> str:
    path = " > ".join(observation.industry_path)
    return f"{observation.source_ordinal:03d} | {path}"


def _observation_to_industry(
    observation: SoiIndustryObservation,
) -> dict[str, object]:
    if observation.form == "sole_proprietorship":
        wage_names = ("payroll",)
        capital_names = ("depreciation_deduction",)
        capital_measure = "depreciation_deduction_flow_over_receipts"
        proxy = True
    elif observation.form == "partnership":
        wage_names = ("cost_labor", "salaries")
        capital_names = ("gross_depreciable_assets",)
        capital_measure = "gross_depreciable_assets_over_receipts"
        proxy = False
    else:
        wage_names = ("officer_compensation", "salaries")
        capital_names = ("gross_depreciable_assets",)
        capital_measure = "gross_depreciable_assets_over_receipts"
        proxy = False

    wage_share, wage_flag, wage_null = _ratio(observation, numerator_names=wage_names)
    ubia_intensity, capital_flag, capital_null = _ratio(
        observation, numerator_names=capital_names
    )
    hint, hint_basis = census_bin_hint(observation.industry_path)
    raw_amounts = {
        "receipts": observation.receipts,
        "salaries": observation.salaries,
        "cost_labor": observation.cost_labor,
        "officer_compensation": observation.officer_compensation,
        "guaranteed_payments_excluded": (observation.guaranteed_payments_excluded),
        "payroll": observation.payroll,
        "gross_depreciable_assets": observation.gross_depreciable_assets,
        "depreciation_deduction": observation.depreciation_deduction,
    }
    return {
        "industry_key": _industry_key(observation),
        "form": observation.form,
        "published_label": observation.published_label,
        "industry_path": list(observation.industry_path),
        "source_ordinal": observation.source_ordinal,
        "industry_level": observation.industry_level,
        "is_aggregate": observation.is_aggregate,
        "census_bin_hint": hint,
        "census_bin_hint_basis": hint_basis,
        "wage_share": wage_share,
        "ubia_intensity": ubia_intensity,
        "proxy": proxy,
        "capital_measure": capital_measure,
        "raw_amounts_thousands": raw_amounts,
        "publication_flags": {
            **dict(observation.publication_flags),
            "wage_share": wage_flag,
            "ubia_intensity": capital_flag,
        },
        "null_reasons": {
            "wage_share": wage_null,
            "ubia_intensity": capital_null,
        },
        "provenance": dict(observation.provenance),
    }


def _range(values: Sequence[float]) -> dict[str, float] | None:
    if not values:
        return None
    return {"minimum": float(min(values)), "maximum": float(max(values))}


def _form_summary(industries: Sequence[Mapping[str, object]]) -> dict[str, object]:
    finest = [
        industry
        for industry in industries
        if not industry["is_aggregate"] and industry["industry_level"] != "unallocable"
    ]
    wages = [
        float(industry["wage_share"])
        for industry in finest
        if industry["wage_share"] is not None
    ]
    capital = [
        float(industry["ubia_intensity"])
        for industry in finest
        if industry["ubia_intensity"] is not None
    ]
    return {
        "published_industry_count": len(industries),
        "finest_industry_count": len(finest),
        "valid_finest_wage_share_count": len(wages),
        "valid_finest_ubia_intensity_count": len(capital),
        "finest_wage_share_range": _range(wages),
        "finest_ubia_intensity_range": _range(capital),
    }


def build_qbi_wage_capital_priors_resource(
    *,
    sole_proprietorship: Sequence[SoiIndustryObservation],
    partnership: Sequence[SoiIndustryObservation],
    s_corporation: Sequence[SoiIndustryObservation],
    all_corporation_review: Mapping[str, object],
    provenance: Mapping[str, object],
) -> dict[str, object]:
    """Build the provisional SOI wage-share and capital-intensity resource."""

    observations = {
        "sole_proprietorship": list(sole_proprietorship),
        "partnership": list(partnership),
        "s_corporation": list(s_corporation),
    }
    for form, form_observations in observations.items():
        if not form_observations:
            raise ValueError(f"SOI {form} observations must be nonempty.")
        if any(observation.form != form for observation in form_observations):
            raise ValueError(f"SOI {form} observations contain another form.")

    industries_by_form = {
        form: [_observation_to_industry(observation) for observation in values]
        for form, values in observations.items()
    }
    payload: dict[str, object] = {
        "schema_version": 1,
        "resource_id": "qbi_wage_capital_priors_v1",
        "provisional": True,
        "purpose": (
            "Estimated public SOI industry priors for total wage expense over "
            "business receipts and gross depreciable assets over receipts. "
            "Simulation wiring and crosswalk adjudication occur later."
        ),
        "forms": {
            "sole_proprietorship": {
                "tax_year": 2023,
                "source_tables": [
                    "23sp01br.xls Table 1",
                    "23sp02is.xls Table 2",
                ],
                "wage_measure": (
                    "Published Payroll (salaries and wages plus cost of labor) "
                    "divided by business receipts"
                ),
                "capital_measure": (
                    "Depreciation deduction including Form 8829 divided by "
                    "business receipts; a flow proxy, not UBIA"
                ),
                "summary": _form_summary(industries_by_form["sole_proprietorship"]),
                "industries": industries_by_form["sole_proprietorship"],
            },
            "partnership": {
                "tax_year": 2023,
                "source_tables": [
                    "23pa01.xlsx Table 1",
                    "23pa03.xlsx Table 3",
                ],
                "wage_measure": (
                    "Cost of labor plus salaries and wages, excluding guaranteed "
                    "payments to partners, divided by business receipts"
                ),
                "capital_measure": (
                    "Gross depreciable assets divided by business receipts; "
                    "closest public book-value analog to UBIA"
                ),
                "summary": _form_summary(industries_by_form["partnership"]),
                "industries": industries_by_form["partnership"],
            },
            "s_corporation": {
                "tax_year": 2022,
                "source_tables": ["22co61ccr.xlsx Table 6.1"],
                "wage_measure": (
                    "Compensation of officers plus salaries and wages divided "
                    "by business receipts"
                ),
                "capital_measure": (
                    "Gross depreciable assets divided by business receipts; "
                    "closest public book-value analog to UBIA"
                ),
                "summary": _form_summary(industries_by_form["s_corporation"]),
                "industries": industries_by_form["s_corporation"],
            },
        },
        "source_exclusions": {
            "all_corporation_minor_industry_table": dict(all_corporation_review)
        },
        "derivation_notes": [
            (
                "All monetary amounts share thousands-of-dollars units, which "
                "cancel in the ratios. Finest-industry summary ranges exclude "
                "published totals, disclosure-combined/deleted values, and "
                "nonpositive receipts; caution estimates remain."
            ),
            (
                "Sole-proprietor Table 1 Payroll is canonical and exactly "
                "cross-checks to Table 2 cost of labor plus salaries and wages. "
                "Table 1 depreciation is used because it includes Form 8829."
            ),
            (
                "Partnership guaranteed payments are recorded as an excluded "
                "diagnostic because partners are not W-2 employees."
            ),
            (
                "The 18 published partnership sectors are labeled sector totals "
                "but treated as finest available because the public table has no "
                "finer partnership industry rows. Unallocable categories are "
                "excluded from summary ranges for every form."
            ),
            (
                "Partnership balance-sheet aggregates exclude some small "
                "partnerships exempt from Schedule L while receipts cover all "
                "partnerships, which can bias capital intensity downward."
            ),
            (
                "S-corporation Table 6.1 publishes no separate COGS labor line, "
                "so wages embedded in cost of goods sold may be omitted."
            ),
            (
                "Gross depreciable assets are pre-accumulated-depreciation book "
                "stocks, not statutory tax UBIA. Land is excluded."
            ),
            (
                "Legacy XLS disclosure and caution markers are read from custom "
                "number formats. Underlying numeric values marked ** are never "
                "treated as industry-specific estimates."
            ),
            (
                "census_bin_hint is a conservative seam to the SCF seven-bin "
                "industry collapse. Mixed published groups remain null rather "
                "than being forced into a bin."
            ),
        ],
        "provenance": dict(provenance),
        "judgment_calls": [
            (
                "The resource preserves published aggregates and finest detail; "
                "summary ranges use only nonaggregate published detail."
            ),
            (
                "Payroll, cost-of-labor inclusion, officer compensation, "
                "guaranteed-payment exclusion, gross-book-assets treatment, and "
                "the conservative SCF-bin hints should be re-adjudicated before "
                "simulation use."
            ),
            (
                "Sole-proprietor and partnership evidence use tax year 2023; "
                "the latest available S-corporation table is tax year 2022."
            ),
        ],
    }
    validate_qbi_wage_capital_priors_resource(payload)
    return payload


def validate_qbi_wage_capital_priors_resource(
    payload: Mapping[str, object],
) -> None:
    """Validate the committed SOI evidence-resource contract."""

    root = _exact_mapping(
        payload,
        {
            "schema_version",
            "resource_id",
            "provisional",
            "purpose",
            "forms",
            "source_exclusions",
            "derivation_notes",
            "provenance",
            "judgment_calls",
        },
        "QBI wage/capital resource",
    )
    if root["schema_version"] != 1:
        raise ValueError("QBI wage/capital schema_version must equal 1.")
    if root["resource_id"] != "qbi_wage_capital_priors_v1":
        raise ValueError("QBI wage/capital resource_id is not recognized.")
    if root["provisional"] is not True:
        raise ValueError("QBI wage/capital resource must remain provisional.")
    _nonempty_text(root["purpose"], "purpose")
    _validate_text_list(root["derivation_notes"], "derivation_notes")
    _validate_text_list(root["judgment_calls"], "judgment_calls")
    _validate_provenance(root["provenance"], "provenance")
    exclusions = _exact_mapping(
        root["source_exclusions"],
        {"all_corporation_minor_industry_table"},
        "source_exclusions",
    )
    all_corporation = _exact_mapping(
        exclusions["all_corporation_minor_industry_table"],
        {"filename", "table", "tax_year", "review_status", "reason"},
        "source_exclusions.all_corporation_minor_industry_table",
    )
    if (
        all_corporation["tax_year"] != 2022
        or all_corporation["review_status"] != "inspected_not_used"
    ):
        raise ValueError("All-corporation exclusion must remain review-only for 2022.")
    _nonempty_text(
        all_corporation["reason"],
        "source_exclusions.all_corporation_minor_industry_table.reason",
    )

    forms = _exact_mapping(root["forms"], set(SOI_ENTITY_FORMS), "forms")
    form_specs = {
        "sole_proprietorship": {
            "tax_year": 2023,
            "proxy": True,
            "capital_measure": "depreciation_deduction_flow_over_receipts",
            "wage_names": ("payroll",),
            "capital_names": ("depreciation_deduction",),
            "provenance_extras": {"wage_component_crosscheck_cells"},
        },
        "partnership": {
            "tax_year": 2023,
            "proxy": False,
            "capital_measure": "gross_depreciable_assets_over_receipts",
            "wage_names": ("cost_labor", "salaries"),
            "capital_names": ("gross_depreciable_assets",),
            "provenance_extras": {
                "excluded_wage_cell",
                "depreciation_deduction_cell",
            },
        },
        "s_corporation": {
            "tax_year": 2022,
            "proxy": False,
            "capital_measure": "gross_depreciable_assets_over_receipts",
            "wage_names": ("officer_compensation", "salaries"),
            "capital_names": ("gross_depreciable_assets",),
            "provenance_extras": {"depreciation_deduction_cell"},
        },
    }
    raw_amount_keys = {
        "receipts",
        "salaries",
        "cost_labor",
        "officer_compensation",
        "guaranteed_payments_excluded",
        "payroll",
        "gross_depreciable_assets",
        "depreciation_deduction",
    }
    publication_flag_keys = raw_amount_keys | {"wage_share", "ubia_intensity"}
    industry_keys = {
        "industry_key",
        "form",
        "published_label",
        "industry_path",
        "source_ordinal",
        "industry_level",
        "is_aggregate",
        "census_bin_hint",
        "census_bin_hint_basis",
        "wage_share",
        "ubia_intensity",
        "proxy",
        "capital_measure",
        "raw_amounts_thousands",
        "publication_flags",
        "null_reasons",
        "provenance",
    }
    form_keys = {
        "tax_year",
        "source_tables",
        "wage_measure",
        "capital_measure",
        "summary",
        "industries",
    }
    summary_keys = {
        "published_industry_count",
        "finest_industry_count",
        "valid_finest_wage_share_count",
        "valid_finest_ubia_intensity_count",
        "finest_wage_share_range",
        "finest_ubia_intensity_range",
    }
    provenance_keys = {
        "source_tables",
        "sheet_names",
        "tax_year",
        "units",
        "industry_cells",
        "receipts_cell",
        "wage_cells",
        "capital_cell",
        "calculation",
    }

    for form in SOI_ENTITY_FORMS:
        spec = form_specs[form]
        form_payload = _exact_mapping(forms[form], form_keys, f"forms.{form}")
        if form_payload["tax_year"] != spec["tax_year"]:
            raise ValueError(f"forms.{form}.tax_year is not recognized.")
        _validate_text_list(
            form_payload["source_tables"], f"forms.{form}.source_tables"
        )
        _nonempty_text(form_payload["wage_measure"], f"forms.{form}.wage_measure")
        _nonempty_text(form_payload["capital_measure"], f"forms.{form}.capital_measure")
        summary = _exact_mapping(
            form_payload["summary"], summary_keys, f"forms.{form}.summary"
        )
        industries = _list(form_payload["industries"], f"forms.{form}.industries")
        if not industries:
            raise ValueError(f"forms.{form}.industries must be nonempty.")
        seen: set[str] = set()
        ordinals: list[int] = []
        for index, raw_industry in enumerate(industries):
            location = f"forms.{form}.industries[{index}]"
            industry = _exact_mapping(raw_industry, industry_keys, location)
            key = _nonempty_text(industry["industry_key"], f"{location}.industry_key")
            if key in seen:
                raise ValueError(f"forms.{form} duplicates industry key {key!r}.")
            seen.add(key)
            if industry["form"] != form:
                raise ValueError(f"Industry {key!r} has the wrong entity form.")
            label = _nonempty_text(
                industry["published_label"], f"{location}.published_label"
            )
            path = _validate_text_list(
                industry["industry_path"], f"{location}.industry_path"
            )
            if path[-1] != label:
                raise ValueError(f"Industry {key!r} label does not match its path.")
            ordinal = _nonnegative_integer(
                industry["source_ordinal"], f"{location}.source_ordinal"
            )
            if ordinal == 0:
                raise ValueError(f"Industry {key!r} source ordinal must be positive.")
            ordinals.append(ordinal)
            if key != f"{ordinal:03d} | {' > '.join(str(item) for item in path)}":
                raise ValueError(f"Industry {key!r} has an inconsistent key.")
            level = industry["industry_level"]
            if level not in {"all", "sector_total", "published_detail", "unallocable"}:
                raise ValueError(f"Industry {key!r} has an unknown level.")
            if not isinstance(industry["is_aggregate"], bool):
                raise ValueError(f"Industry {key!r} is_aggregate must be boolean.")
            if (level == "all") != industry["is_aggregate"] and level != "sector_total":
                raise ValueError(
                    f"Industry {key!r} has inconsistent aggregate metadata."
                )
            hint = industry["census_bin_hint"]
            if hint is not None and (
                isinstance(hint, bool)
                or not isinstance(hint, int)
                or hint not in range(1, 8)
            ):
                raise ValueError(f"Industry {key!r} has invalid SCF-bin hint.")
            _nonempty_text(
                industry["census_bin_hint_basis"],
                f"{location}.census_bin_hint_basis",
            )
            if industry["proxy"] is not spec["proxy"]:
                raise ValueError(f"Industry {key!r} has the wrong proxy flag.")
            if industry["capital_measure"] != spec["capital_measure"]:
                raise ValueError(f"Industry {key!r} has the wrong capital measure.")

            for measure in ("wage_share", "ubia_intensity"):
                value = industry[measure]
                if value is None:
                    continue
                _finite_number(value, f"{location}.{measure}", minimum=0.0)

            raw_amounts = _exact_mapping(
                industry["raw_amounts_thousands"],
                raw_amount_keys,
                f"{location}.raw_amounts_thousands",
            )
            for name, value in raw_amounts.items():
                if value is not None:
                    _finite_number(value, f"{location}.raw_amounts_thousands.{name}")
            flags = _exact_mapping(
                industry["publication_flags"],
                publication_flag_keys,
                f"{location}.publication_flags",
            )
            unknown_flags = set(flags.values()) - SOI_PUBLICATION_FLAGS
            if unknown_flags:
                raise ValueError(
                    f"Industry {key!r} has unknown publication flags "
                    f"{sorted(unknown_flags)}."
                )
            null_reasons = _exact_mapping(
                industry["null_reasons"],
                {"wage_share", "ubia_intensity"},
                f"{location}.null_reasons",
            )
            ratio_specs = {
                "wage_share": spec["wage_names"],
                "ubia_intensity": spec["capital_names"],
            }
            for measure, names in ratio_specs.items():
                ratio_flags = [str(flags["receipts"])]
                numerator = 0.0
                expected_null: str | None = None
                for name in names:
                    ratio_flags.append(str(flags[name]))
                    value = raw_amounts[name]
                    if value is None:
                        expected_null = f"{name}_not_published"
                        break
                    numerator += float(value)
                receipts = raw_amounts["receipts"]
                if expected_null is None:
                    if receipts is None:
                        expected_null = "receipts_not_published"
                    elif float(receipts) <= 0.0:
                        expected_null = "nonpositive_receipts"
                expected_flag = _combined_flag(ratio_flags)
                if flags[measure] != expected_flag:
                    raise ValueError(
                        f"Industry {key!r} {measure} publication flag is inconsistent."
                    )
                if null_reasons[measure] != expected_null:
                    raise ValueError(
                        f"Industry {key!r} {measure} null reason is inconsistent."
                    )
                ratio = industry[measure]
                if expected_null is not None:
                    if ratio is not None:
                        raise ValueError(
                            f"Industry {key!r} {measure} must be null for its inputs."
                        )
                else:
                    expected_ratio = numerator / float(receipts)
                    if ratio is None or not math.isclose(
                        float(ratio),
                        expected_ratio,
                        rel_tol=0.0,
                        abs_tol=1e-12,
                    ):
                        raise ValueError(
                            f"Industry {key!r} {measure} disagrees with raw amounts."
                        )

            expected_provenance_keys = provenance_keys | set(spec["provenance_extras"])
            provenance = _exact_mapping(
                industry["provenance"],
                expected_provenance_keys,
                f"{location}.provenance",
            )
            if provenance["tax_year"] != spec["tax_year"]:
                raise ValueError(f"Industry {key!r} provenance has the wrong year.")
            if provenance["units"] != "thousands_of_dollars":
                raise ValueError(f"Industry {key!r} provenance has the wrong units.")
            for field in (
                "source_tables",
                "sheet_names",
                "industry_cells",
                "wage_cells",
            ):
                _validate_text_list(provenance[field], f"{location}.provenance.{field}")
            for field in (
                "receipts_cell",
                "capital_cell",
                *spec["provenance_extras"],
            ):
                value = provenance[field]
                if isinstance(value, list):
                    _validate_text_list(value, f"{location}.provenance.{field}")
                else:
                    _nonempty_text(value, f"{location}.provenance.{field}")
            calculation = _exact_mapping(
                provenance["calculation"],
                {"wage_share", "ubia_intensity"},
                f"{location}.provenance.calculation",
            )
            _nonempty_text(
                calculation["wage_share"],
                f"{location}.provenance.calculation.wage_share",
            )
            _nonempty_text(
                calculation["ubia_intensity"],
                f"{location}.provenance.calculation.ubia_intensity",
            )
        if ordinals != list(range(1, len(industries) + 1)):
            raise ValueError(f"forms.{form} source ordinals must be contiguous.")
        if dict(summary) != _form_summary(industries):
            raise ValueError(f"forms.{form}.summary disagrees with its industries.")
