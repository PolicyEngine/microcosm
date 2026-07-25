"""Synthetic-fixture and packaged-resource tests for QBI v3 evidence."""

from __future__ import annotations

import importlib.util
import json
import zipfile
from copy import deepcopy
from importlib.resources import files
from pathlib import Path

import numpy as np
import pandas as pd
import pytest
from openpyxl import Workbook

from populace.build.us_runtime.qbi_v3_evidence import (
    SCF_MINIMUM_UNWEIGHTED_N,
    SoiIndustryObservation,
    build_qbi_employer_structure_resource,
    build_qbi_wage_capital_priors_resource,
    build_scf_business_records,
    census_bin_hint,
    parse_partnership_soi_workbooks,
    parse_s_corporation_soi_workbook,
    validate_qbi_employer_structure_resource,
    validate_qbi_wage_capital_priors_resource,
    weighted_inverse_cdf,
)

ROOT = Path(__file__).resolve().parents[3]
BUILDER_PATH = ROOT / "tools/build_us_qbi_v3_evidence.py"
RESOURCE_NAMES = {
    "qbi_employer_structure_v1.json",
    "qbi_wage_capital_priors_v1.json",
}


def _load_builder():
    spec = importlib.util.spec_from_file_location(
        "qbi_v3_evidence_builder", BUILDER_PATH
    )
    assert spec is not None and spec.loader is not None
    builder = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(builder)
    return builder


def _scf_fixture(
    businesses: list[dict[str, float | int]],
) -> pd.DataFrame:
    rows: list[dict[str, float | int]] = []
    for household_id, business in enumerate(businesses, start=1):
        business_count = int(business.get("business_count", 1))
        for implicate in range(1, 6):
            row: dict[str, float | int] = {
                "y1": household_id * 10 + implicate,
                "x42001": float(business.get("weight", 100.0)),
                "x3103": 1,
                "x3104": 1,
                "x3105": business_count,
                "x3107": int(business.get("industry", 1)),
                "x3111": int(business.get("headcount", 1)),
                "x3113": 1,
                "x3114": 5,
                "x3119": int(business.get("legal_form", 2)),
                "x3128": int(business.get("ownership", 10_000)),
                "x3131": float(business.get("receipts", 20_000.0)),
                "x3132": float(business.get("net_income", 10_000.0)),
                "x3207": int(business.get("second_industry", 6)),
                "x3211": int(business.get("second_headcount", 1)),
                "x3213": 1,
                "x3214": 5,
                "x3219": int(business.get("second_legal_form", 11)),
                "x3228": int(business.get("second_ownership", 5_000)),
                "x3231": float(business.get("second_receipts", 40_000.0)),
                "x3232": float(business.get("second_net_income", 8_000.0)),
            }
            rows.append(row)
    return pd.DataFrame.from_records(rows)


def _provenance() -> dict[str, object]:
    return {
        "generated_by": "synthetic test",
        "run_command": "synthetic fixture",
        "inputs": [{"filename": "synthetic.dta", "sha256": "0" * 64}],
    }


def _employer_cell(
    payload: dict[str, object],
    *,
    income_band: str,
    legal_form_group: str,
    industry_code: int,
) -> dict[str, object]:
    cells = payload["cells"]
    assert isinstance(cells, list)
    return next(
        cell
        for cell in cells
        if cell["income_band"] == income_band
        and cell["legal_form_group"] == legal_form_group
        and cell["industry_code"] == industry_code
    )


def test_scf_implicates_pool_with_weight_divided_by_five() -> None:
    source = _scf_fixture(
        [
            {"weight": 100.0, "net_income": -1.0, "receipts": -1.0},
            {
                "weight": 200.0,
                "business_count": 2,
                "second_ownership": 2_500,
                "second_net_income": 8_000.0,
            },
        ]
    )

    records = build_scf_business_records(source)

    assert len(records) == 15
    assert records["weight"].sum() == pytest.approx(500.0)
    assert records.loc[records["business_slot"].eq(1), "weight"].sum() == 300.0
    first = records.loc[records["household_id"].eq(1)].iloc[0]
    assert first["gross_receipts"] == 0.0
    assert first["whole_net_income"] == 0.0
    second = records.loc[records["business_slot"].eq(2)].iloc[0]
    assert second["owned_net_income"] == 2_000.0


def test_builder_reads_equivalent_scf_dta_and_zip_fixtures(tmp_path) -> None:
    source = _scf_fixture([{"weight": 100.0}, {"weight": 200.0, "business_count": 2}])
    dta_path = tmp_path / "p22i6.dta"
    zip_path = tmp_path / "scf2022s.zip"
    source.to_stata(dta_path, write_index=False)
    with zipfile.ZipFile(zip_path, "w") as archive:
        archive.write(dta_path, arcname="p22i6.dta")
    builder = _load_builder()

    dta_frame, dta_inputs = builder.read_scf_source(dta_path)
    zip_frame, zip_inputs = builder.read_scf_source(zip_path)

    pd.testing.assert_frame_equal(dta_frame, zip_frame)
    assert [record["role"] for record in dta_inputs] == ["scf_2022_source"]
    assert [record["role"] for record in zip_inputs] == [
        "scf_2022_source",
        "scf_2022_archive_member",
    ]
    assert zip_inputs[1]["sha256"] == dta_inputs[0]["sha256"]


def test_scf_thin_cells_follow_independent_nested_fallbacks() -> None:
    businesses = [
        {
            "industry": 1,
            "headcount": 1 if index < 5 else 3,
            "receipts": 20_000.0,
            "net_income": 10_000.0,
        }
        for index in range(40)
    ]
    businesses.extend(
        {
            "industry": 2,
            "headcount": 6,
            "receipts": 50_000.0,
            "net_income": 10_000.0,
        }
        for _ in range(10)
    )

    payload = build_qbi_employer_structure_resource(
        _scf_fixture(businesses),
        provenance=_provenance(),
        minimum_unweighted_n=30.0,
    )

    assert payload["source"]["variables"]["active_management_screeners"] == [
        "X3103",
        "X3104",
    ]
    assert "do not further restrict" in payload["source"]["record_selection"]
    comparisons = payload["external_anchor"]["scf_comparison"]
    assert comparisons["main_proxy_including_informal_code_40"]["legal_form_codes"] == [
        1,
        2,
        3,
        11,
        40,
    ]
    assert (
        "Owned net income is greater than zero"
        in comparisons["strict_form_sensitivity_excluding_code_40"]["sample_definition"]
    )

    exact = _employer_cell(
        payload,
        income_band="0_to_25k",
        legal_form_group="sole_or_informal",
        industry_code=1,
    )
    assert exact["requested_counts"]["implicate_adjusted_unweighted_n"] == 40.0
    assert exact["employer_presence"]["estimate_level"] == "exact"
    assert exact["employer_presence"]["probability_headcount_gt_1"] == pytest.approx(
        0.875
    )
    assert exact["headcount_size_distribution"]["estimate_level"] == "exact"

    thin = _employer_cell(
        payload,
        income_band="0_to_25k",
        legal_form_group="sole_or_informal",
        industry_code=2,
    )
    assert thin["requested_counts"]["implicate_adjusted_unweighted_n"] == 10.0
    assert thin["employer_presence"]["estimate_level"] == "income_form"
    assert thin["headcount_size_distribution"]["estimate_level"] == "income_form"

    empty_income = _employer_cell(
        payload,
        income_band="over_1m",
        legal_form_group="sole_or_informal",
        industry_code=7,
    )
    assert empty_income["employer_presence"]["estimate_level"] == "form"
    assert empty_income["headcount_size_distribution"]["estimate_level"] == "form"
    assert sum(
        empty_income["headcount_size_distribution"]["shares"].values()
    ) == pytest.approx(1.0)


def test_weighted_profit_margin_quantiles_use_inverse_cdf() -> None:
    quantiles = weighted_inverse_cdf(
        np.array([0.1, 0.2, 0.9]),
        np.array([1.0, 3.0, 1.0]),
        (0.05, 0.25, 0.5, 0.75, 0.95),
    )

    assert quantiles == {
        "q05": 0.1,
        "q25": 0.2,
        "q50": 0.2,
        "q75": 0.2,
        "q95": 0.9,
    }


def test_employer_resource_schema_rejects_nonprovisional_payload() -> None:
    source = _scf_fixture(
        [
            {
                "industry": (index % 6) + 1,
                "headcount": 3,
                "receipts": 20_000.0,
                "net_income": 10_000.0,
            }
            for index in range(35)
        ]
    )
    payload = build_qbi_employer_structure_resource(
        source,
        provenance=_provenance(),
        minimum_unweighted_n=SCF_MINIMUM_UNWEIGHTED_N,
    )
    broken = deepcopy(payload)
    broken["provisional"] = False

    with pytest.raises(ValueError, match="must remain provisional"):
        validate_qbi_employer_structure_resource(broken)


def _publication_flags(**overrides: str) -> dict[str, str]:
    flags = {
        "receipts": "published",
        "salaries": "published",
        "cost_labor": "published",
        "officer_compensation": "published",
        "guaranteed_payments_excluded": "published",
        "payroll": "published",
        "gross_depreciable_assets": "published",
        "depreciation_deduction": "published",
    }
    flags.update(overrides)
    return flags


def _soi_observation(
    *,
    form: str,
    label: str,
    ordinal: int,
    receipts: float | None = 1_000.0,
    salaries: float | None = 100.0,
    cost_labor: float | None = 50.0,
    officer_compensation: float | None = 200.0,
    guaranteed_payments: float | None = 900.0,
    payroll: float | None = 125.0,
    assets: float | None = 500.0,
    depreciation: float | None = 40.0,
    flags: dict[str, str] | None = None,
    aggregate: bool = False,
) -> SoiIndustryObservation:
    return SoiIndustryObservation(
        form=form,
        tax_year=2022 if form == "s_corporation" else 2023,
        published_label=label,
        industry_path=(label,),
        source_ordinal=ordinal,
        industry_level="sector_total" if aggregate else "published_detail",
        is_aggregate=aggregate,
        receipts=receipts,
        salaries=salaries,
        cost_labor=cost_labor,
        officer_compensation=officer_compensation,
        guaranteed_payments_excluded=guaranteed_payments,
        payroll=payroll,
        gross_depreciable_assets=assets,
        depreciation_deduction=depreciation,
        publication_flags=flags or _publication_flags(),
        provenance={
            "source_tables": ["synthetic"],
            "tax_year": 2023,
            "units": "thousands_of_dollars",
            "industry_cells": ["Sheet1!B4"],
            "receipts_cell": "Sheet1!B18",
            "wage_cells": ["Sheet1!B26"],
            "capital_cell": "Sheet1!B29",
            "calculation": {"wage_share": "synthetic"},
        },
    )


def test_soi_ratio_builder_uses_form_specific_numerators_and_proxy_flag() -> None:
    payload = build_qbi_wage_capital_priors_resource(
        sole_proprietorship=[
            _soi_observation(
                form="sole_proprietorship",
                label="Taxi and limousine service",
                ordinal=1,
            )
        ],
        partnership=[
            _soi_observation(
                form="partnership",
                label="Construction",
                ordinal=1,
            )
        ],
        s_corporation=[
            _soi_observation(
                form="s_corporation",
                label="Manufacturing",
                ordinal=1,
            )
        ],
        all_corporation_review={
            "filename": "synthetic.xlsx",
            "review_status": "inspected_not_used",
        },
        provenance=_provenance(),
    )

    forms = payload["forms"]
    sole = forms["sole_proprietorship"]["industries"][0]
    partnership = forms["partnership"]["industries"][0]
    s_corporation = forms["s_corporation"]["industries"][0]
    assert sole["wage_share"] == pytest.approx(0.125)
    assert sole["ubia_intensity"] == pytest.approx(0.04)
    assert sole["proxy"] is True
    assert partnership["wage_share"] == pytest.approx(0.15)
    assert partnership["ubia_intensity"] == pytest.approx(0.5)
    assert partnership["raw_amounts_thousands"]["guaranteed_payments_excluded"] == 900.0
    assert s_corporation["wage_share"] == pytest.approx(0.3)
    assert s_corporation["ubia_intensity"] == pytest.approx(0.5)
    assert s_corporation["proxy"] is False


def _write_partnership_fixtures(tmp_path):
    income_path = tmp_path / "partnership-income.xlsx"
    balance_path = tmp_path / "partnership-balance.xlsx"
    income_book = Workbook()
    balance_book = Workbook()
    income = income_book.active
    balance = balance_book.active
    income.title = "Sheet1"
    balance.title = "Sheet1"
    income["A1"] = "Table 1, Tax Year 2023"
    balance["A1"] = "Table 3, Tax Year 2023"
    labels = [
        "All industries",
        *[f"Sector {index}" for index in range(1, 19)],
        "Unallocable",
    ]
    for column, label in enumerate(labels, start=2):
        income.cell(4, column, label)
        balance.cell(4, column, label)
        income.cell(18, column, 1_000.0)
        income.cell(26, column, 50.0)
        income.cell(30, column, 100.0)
        income.cell(31, column, 900.0)
        income.cell(37, column, 40.0)
        balance.cell(29, column, 500.0)
    income.cell(18, 21, 0.0)
    balance.cell(29, 21, 0.0)
    income_book.save(income_path)
    balance_book.save(balance_path)
    return income_path, balance_path


def _write_s_corporation_fixture(tmp_path):
    path = tmp_path / "s-corporation.xlsx"
    book = Workbook()
    sheet = book.active
    sheet.title = "Table 6.1"
    sheet["A2"] = "Table 6.1 by Major Industry, Tax Year 2022"
    sheet.merge_cells("B5:B7")
    sheet["B5"] = "All industries"
    sheet.merge_cells("C5:D5")
    sheet["C5"] = "Construction"
    sheet.merge_cells("C6:C7")
    sheet["C6"] = "Total"
    sheet.merge_cells("D6:D7")
    sheet["D6"] = "Construction of buildings"
    sheet.merge_cells("E5:E7")
    sheet["E5"] = "Insurance carriers"
    for ordinal, column in enumerate(range(2, 6), start=1):
        sheet.cell(8, column, ordinal)
        sheet.cell(21, column, 500.0)
        sheet.cell(43, column, 1_000.0)
        sheet.cell(49, column, 200.0)
        sheet.cell(50, column, 100.0)
        sheet.cell(57, column, 40.0)
    sheet["E43"] = "d"
    sheet["E49"] = "d"
    book.save(path)
    return path


def test_soi_xlsx_parsers_use_published_cells_and_disclosure_deletions(
    tmp_path,
) -> None:
    partnership_paths = _write_partnership_fixtures(tmp_path)
    s_corporation_path = _write_s_corporation_fixture(tmp_path)

    partnership = parse_partnership_soi_workbooks(*partnership_paths)
    s_corporation = parse_s_corporation_soi_workbook(s_corporation_path)

    assert len(partnership) == 20
    assert partnership[1].receipts == 1_000.0
    assert partnership[1].industry_level == "sector_total"
    assert partnership[1].guaranteed_payments_excluded == 900.0
    assert (
        partnership[1].provenance["depreciation_deduction_cell"].endswith("Sheet1!C37")
    )
    assert partnership[-1].industry_level == "unallocable"
    assert len(s_corporation) == 4
    assert s_corporation[1].is_aggregate is True
    assert s_corporation[2].industry_path == (
        "Construction",
        "Construction of buildings",
    )
    assert s_corporation[3].receipts is None
    assert s_corporation[3].publication_flags["receipts"] == "deleted_for_disclosure"
    assert s_corporation[1].provenance["depreciation_deduction_cell"] == "Table 6.1!C57"


@pytest.mark.parametrize(
    ("path", "expected"),
    [
        (("Manufacturing", "Machinery manufacturing"), 3),
        (("Information", "Software publishing"), 5),
        (("Accommodation and food services",), None),
        (("Professional services", "Veterinary services"), 1),
    ],
)
def test_census_bin_hints_are_conservative(
    path: tuple[str, ...],
    expected: int | None,
) -> None:
    assert census_bin_hint(path)[0] == expected


def test_wage_capital_schema_rejects_negative_ratio() -> None:
    payload = build_qbi_wage_capital_priors_resource(
        sole_proprietorship=[
            _soi_observation(
                form="sole_proprietorship",
                label="Retail trade",
                ordinal=1,
            )
        ],
        partnership=[
            _soi_observation(
                form="partnership",
                label="Retail trade",
                ordinal=1,
            )
        ],
        s_corporation=[
            _soi_observation(
                form="s_corporation",
                label="Retail trade",
                ordinal=1,
            )
        ],
        all_corporation_review={
            "filename": "synthetic.xlsx",
            "review_status": "inspected_not_used",
        },
        provenance=_provenance(),
    )
    broken = deepcopy(payload)
    broken["forms"]["partnership"]["industries"][0]["wage_share"] = -0.1

    with pytest.raises(ValueError, match="finite and nonnegative"):
        validate_qbi_wage_capital_priors_resource(broken)


def test_packaged_qbi_v3_resources_validate_and_remain_provisional() -> None:
    package = files("populace.build.us")
    employer = json.loads(
        package.joinpath("qbi_employer_structure_v1.json").read_text(encoding="utf-8")
    )
    wage_capital = json.loads(
        package.joinpath("qbi_wage_capital_priors_v1.json").read_text(encoding="utf-8")
    )

    validate_qbi_employer_structure_resource(employer)
    validate_qbi_wage_capital_priors_resource(wage_capital)
    assert employer["provisional"] is True
    assert wage_capital["provisional"] is True
    sole_industries = wage_capital["forms"]["sole_proprietorship"]["industries"]
    unclassified = next(
        industry
        for industry in sole_industries
        if industry["published_label"] == "Unclassified establishments"
    )
    assert unclassified["industry_level"] == "unallocable"


def test_qbi_v3_resources_are_declared_as_specs_without_entrypoints() -> None:
    package = files("populace.build.us")
    manifest = json.loads(
        package.joinpath("country_package.json").read_text(encoding="utf-8")
    )

    assert RESOURCE_NAMES.issubset(manifest["resources"])
    for resource_name in RESOURCE_NAMES:
        assert resource_name.endswith(".json")
        rendered = package.joinpath(resource_name).read_text(encoding="utf-8")
        assert ".py:" not in rendered
